import streamlit as st
import pandas as pd

from datetime import date, time
from io import BytesIO
from html import escape
import math
import re


# ============================================================
# PDF
# ============================================================

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as RLImage,
    PageBreak,
)


# ============================================================
# EXCEL
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÃO
# ============================================================

st.set_page_config(
    page_title="Site Engineer Tool",
    page_icon="🏗️",
    layout="wide",
)


# ============================================================
# SESSION STATE
# ============================================================

defaults = {
    "top_block_points": [],
    "dimension_checks": [],
    "floor_slab_points": [],

    "foundation_checklists": {},
    "foundation_bottom_points": [],
    "foundation_top_points": [],

    "daily_reports": {},
    "daily_manpower": [],
    "daily_equipment": [],
    "daily_activities": [],
    "daily_deliveries": [],
    "daily_inspections": [],
    "daily_issues": [],
    "daily_photos": [],
}

for key, value in defaults.items():

    if key not in st.session_state:

        if isinstance(value, list):
            st.session_state[key] = []

        elif isinstance(value, dict):
            st.session_state[key] = {}

        else:
            st.session_state[key] = value


# ============================================================
# FUNÇÕES GERAIS
# ============================================================

def difference_mm(measured, design):

    return (
        measured
        - design
    ) * 1000


def check_tolerance(
    value_mm,
    tolerance_mm,
):

    if abs(value_mm) <= tolerance_mm:
        return "PASS"

    return "NO PASS"


def calculate_diagonal(
    length,
    width,
):

    if (
        length <= 0
        or width <= 0
    ):
        return None

    return math.sqrt(
        (length ** 2)
        + (width ** 2)
    )


def status_icon(status):

    icons = {
        "PASS": "✅ PASS",
        "NO PASS": "⚠️ NO PASS",

        "Completed": "✅ Completed",
        "Complete": "✅ Complete",

        "In Progress": "🟡 In Progress",
        "Not Started": "⚪ Not Started",

        "Yes": "✅ Yes",
        "No": "❌ No",

        "Approved": "✅ Approved",
        "Not Approved": "⚠️ Not Approved",

        "Pending": "🟡 Pending",
        "Draft": "🟡 Draft",

        "N/A": "➖ N/A",
    }

    return icons.get(
        status,
        status,
    )


def safe_index(
    options,
    value,
    default_value=None,
):

    if value in options:
        return options.index(value)

    if default_value in options:
        return options.index(
            default_value
        )

    return 0


def widget_suffix(text):

    text = re.sub(
        r"[^A-Za-z0-9_]+",
        "_",
        str(text),
    )

    text = text.strip("_")

    if not text:
        return "ITEM"

    return text


def format_date_for_display(value):

    if isinstance(
        value,
        date,
    ):

        return value.strftime(
            "%d/%m/%Y"
        )

    if value in [
        None,
        "",
    ]:

        return "-"

    return str(value)


def format_time_for_display(value):

    if isinstance(
        value,
        time,
    ):

        return value.strftime(
            "%H:%M"
        )

    if value in [
        None,
        "",
    ]:

        return "-"

    return str(value)


# ============================================================
# FOUNDATION FUNCTIONS
# ============================================================

def foundation_key(
    project,
    plot,
):

    project_clean = (
        str(project)
        .strip()
        .upper()
    )

    plot_clean = (
        str(plot)
        .strip()
        .upper()
    )

    return (
        f"{project_clean} | "
        f"{plot_clean}"
    )


def calculate_foundation_progress(record):

    if not record:
        return 0

    checks = []

    process_fields = [
        "Excavation Position Setting Out",
        "Excavation Level Setting Out",
        "Excavation Completed",
        "Bottom Survey Completed",
        "Concrete Chairs Installed",
        "Mesh Installed",
        "Concrete Ordered",
        "Concrete Delivery Confirmed",
        "Concrete Pour Completed",
        "Top Survey Completed",
    ]

    for field in process_fields:

        value = record.get(
            field,
            "Not Started",
        )

        checks.append(
            value in [
                "Completed",
                "N/A",
            ]
        )

    yes_no_fields = [
        "Foundation Step Correct",
        "No Standing Water",
        "Excavation Bottom Clean",
        "No Loose Material",
        "Overlap Checked",
        "Cover Checked",
        "AKM Inspection Requested",
        "AKM Inspection Completed",
    ]

    for field in yes_no_fields:

        value = record.get(
            field,
            "No",
        )

        checks.append(
            value in [
                "Yes",
                "N/A",
            ]
        )

    akm_result = record.get(
        "AKM Result",
        "Pending",
    )

    checks.append(
        akm_result in [
            "Approved",
            "N/A",
        ]
    )

    if not checks:
        return 0

    return round(
        (
            sum(checks)
            / len(checks)
        )
        * 100
    )


def foundation_review_items(
    record,
    bottom_points,
    top_points,
):

    review_items = []

    if not record:
        return review_items

    if (
        record.get(
            "Foundation Step Required"
        )
        == "Yes"
        and
        record.get(
            "Foundation Step Correct"
        )
        == "No"
    ):

        review_items.append(
            "Foundation step not correct"
        )

    if (
        record.get(
            "No Standing Water"
        )
        == "No"
    ):

        review_items.append(
            "Standing water in foundation"
        )

    if (
        record.get(
            "Excavation Bottom Clean"
        )
        == "No"
    ):

        review_items.append(
            "Excavation bottom not clean"
        )

    if (
        record.get(
            "No Loose Material"
        )
        == "No"
    ):

        review_items.append(
            "Loose / fallen material present"
        )

    if (
        record.get(
            "Overlap Checked"
        )
        == "No"
    ):

        review_items.append(
            "Mesh overlap not correct / not verified"
        )

    if (
        record.get(
            "Cover Checked"
        )
        == "No"
    ):

        review_items.append(
            "Reinforcement cover not correct / not verified"
        )

    if (
        record.get(
            "AKM Result"
        )
        == "Not Approved"
    ):

        review_items.append(
            "AKM inspection not approved"
        )

    for point in bottom_points:

        if (
            point.get(
                "Status"
            )
            == "NO PASS"
        ):

            review_items.append(
                (
                    "Bottom of Foundation Survey - "
                    f'{point.get("Point", "Point")} '
                    f'({point.get("Difference (mm)", 0):+.1f} mm)'
                )
            )

    for point in top_points:

        if (
            point.get(
                "Status"
            )
            == "NO PASS"
        ):

            review_items.append(
                (
                    "Top of Foundation Survey - "
                    f'{point.get("Point", "Point")} '
                    f'({point.get("Difference (mm)", 0):+.1f} mm)'
                )
            )

    return review_items


def foundation_overall_status(
    record,
    bottom_points,
    top_points,
):

    if not record:
        return "NOT STARTED"

    review_items = foundation_review_items(
        record,
        bottom_points,
        top_points,
    )

    if review_items:
        return "REVIEW REQUIRED"

    concrete_complete = (
        record.get(
            "Concrete Pour Completed"
        )
        == "Completed"
    )

    top_complete = (
        record.get(
            "Top Survey Completed"
        )
        == "Completed"
    )

    akm_approved = (
        record.get(
            "AKM Result"
        )
        in [
            "Approved",
            "N/A",
        ]
    )

    if (
        concrete_complete
        and top_complete
        and akm_approved
    ):

        return "COMPLETE"

    return "IN PROGRESS"


# ============================================================
# DAILY REPORT FUNCTIONS
# ============================================================

def daily_report_key(
    project,
    report_date,
):

    project_clean = (
        str(project)
        .strip()
        .upper()
    )

    if isinstance(
        report_date,
        date,
    ):

        date_clean = (
            report_date
            .isoformat()
        )

    else:

        date_clean = str(
            report_date
        )

    return (
        f"{project_clean} | "
        f"{date_clean}"
    )


def daily_items_for_key(
    collection,
    report_key,
):

    return [
        item
        for item
        in collection
        if (
            item.get(
                "Daily Report Key"
            )
            == report_key
        )
    ]


def calculate_total_manpower(
    report_key,
):

    items = daily_items_for_key(
        st.session_state.daily_manpower,
        report_key,
    )

    total = 0

    for item in items:

        total += int(
            item.get(
                "Quantity",
                0,
            )
        )

    return total


def calculate_total_equipment(
    report_key,
):

    items = daily_items_for_key(
        st.session_state.daily_equipment,
        report_key,
    )

    total = 0

    for item in items:

        total += int(
            item.get(
                "Quantity",
                0,
            )
        )

    return total


# ============================================================
# PDF HELPERS
# ============================================================

def pdf_text(value):

    if value in [
        None,
        "",
    ]:

        return "-"

    return escape(
        str(value)
    )


def make_pdf_paragraph(
    value,
    style,
):

    return Paragraph(
        pdf_text(value),
        style,
    )


def add_pdf_page_number(
    canvas,
    doc,
):

    canvas.saveState()

    canvas.setStrokeColor(
        colors.HexColor(
            "#D9DEE5"
        )
    )

    canvas.line(
        18 * mm,
        14 * mm,
        192 * mm,
        14 * mm,
    )

    canvas.setFont(
        "Helvetica",
        8,
    )

    canvas.setFillColor(
        colors.HexColor(
            "#6B7280"
        )
    )

    canvas.drawString(
        18 * mm,
        8 * mm,
        "GLENVEAGH | Site Daily Report",
    )

    canvas.drawRightString(
        192 * mm,
        8 * mm,
        f"Page {doc.page}",
    )

    canvas.restoreState()


def generate_daily_pdf(
    report_key,
):

    report = (
        st.session_state.daily_reports.get(
            report_key,
            {},
        )
    )

    manpower = daily_items_for_key(
        st.session_state.daily_manpower,
        report_key,
    )

    equipment = daily_items_for_key(
        st.session_state.daily_equipment,
        report_key,
    )

    activities = daily_items_for_key(
        st.session_state.daily_activities,
        report_key,
    )

    deliveries = daily_items_for_key(
        st.session_state.daily_deliveries,
        report_key,
    )

    inspections = daily_items_for_key(
        st.session_state.daily_inspections,
        report_key,
    )

    issues = daily_items_for_key(
        st.session_state.daily_issues,
        report_key,
    )

    photos = daily_items_for_key(
        st.session_state.daily_photos,
        report_key,
    )

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=20 * mm,
        title="GLENVEAGH Site Daily Report",
        author="Site Engineer Tool",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=21,
        leading=24,
        textColor=colors.HexColor(
            "#1F2937"
        ),
        spaceAfter=2 * mm,
    )

    company_style = ParagraphStyle(
        "Company",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        textColor=colors.HexColor(
            "#D15C26"
        ),
    )

    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.HexColor(
            "#6B7280"
        ),
        spaceAfter=4 * mm,
    )

    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        textColor=colors.HexColor(
            "#1F2937"
        ),
        spaceBefore=4 * mm,
        spaceAfter=2.5 * mm,
    )

    normal_style = ParagraphStyle(
        "NormalSmall",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor(
            "#374151"
        ),
    )

    small_style = ParagraphStyle(
        "Small",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor(
            "#4B5563"
        ),
    )

    photo_caption_style = ParagraphStyle(
        "PhotoCaption",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor(
            "#374151"
        ),
        alignment=TA_CENTER,
    )

    # --------------------------------------------------------
    # GLENVEAGH HEADER STYLES
    # Somente o cabeçalho usa as cores da Glenveagh.
    # Todo o restante do relatório permanece inalterado.
    # --------------------------------------------------------

    glenveagh_header_name_style = ParagraphStyle(
        "GlenveaghHeaderName",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=22,
        leading=23,
        textColor=colors.white,
        spaceAfter=0,
    )

    glenveagh_header_tagline_style = ParagraphStyle(
        "GlenveaghHeaderTagline",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=10,
        textColor=colors.HexColor(
            "#B9A323"
        ),
        spaceBefore=1,
    )

    glenveagh_header_title_style = ParagraphStyle(
        "GlenveaghHeaderTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=19,
        textColor=colors.white,
        alignment=2,
        spaceAfter=1,
    )

    glenveagh_header_subtitle_style = ParagraphStyle(
        "GlenveaghHeaderSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=2,
    )

    story = []


    # --------------------------------------------------------
    # HEADER - GLENVEAGH
    # --------------------------------------------------------

    glenveagh_logo_block = [
        Paragraph(
            "Glenveagh",
            glenveagh_header_name_style,
        ),
        Paragraph(
            "Home of the new.",
            glenveagh_header_tagline_style,
        ),
    ]

    glenveagh_report_block = [
        Paragraph(
            "SITE DAILY REPORT",
            glenveagh_header_title_style,
        ),
        Paragraph(
            "Construction &amp; Site Engineering Record",
            glenveagh_header_subtitle_style,
        ),
    ]

    header_table = Table(
        [
            [
                glenveagh_logo_block,
                glenveagh_report_block,
            ]
        ],
        colWidths=[
            78 * mm,
            92 * mm,
        ],
    )

    header_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    colors.HexColor(
                        "#08284A"
                    ),
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),

                (
                    "ALIGN",
                    (1, 0),
                    (1, 0),
                    "RIGHT",
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    10,
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    10,
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    10,
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    10,
                ),

                (
                    "LINEBELOW",
                    (0, 0),
                    (-1, -1),
                    2,
                    colors.HexColor(
                        "#B9A323"
                    ),
                ),
            ]
        )
    )

    story.append(
        header_table
    )

    story.append(
        Spacer(
            1,
            3 * mm,
        )
    )

    story.append(
        Paragraph(
            (
                "Daily site record generated by "
                "Site Engineer Tool"
            ),
            subtitle_style,
        )
    )


    # --------------------------------------------------------
    # REPORT INFO
    # --------------------------------------------------------

    project = report.get(
        "Project",
        "",
    )

    report_date = (
        format_date_for_display(
            report.get(
                "Date",
                "",
            )
        )
    )

    start_time = (
        format_time_for_display(
            report.get(
                "Working Start",
                "",
            )
        )
    )

    finish_time = (
        format_time_for_display(
            report.get(
                "Working Finish",
                "",
            )
        )
    )

    report_status = report.get(
        "Status",
        "Draft",
    )

    info_data = [
        [
            Paragraph(
                "<b>Project / Site</b>",
                small_style,
            ),
            make_pdf_paragraph(
                project,
                normal_style,
            ),
            Paragraph(
                "<b>Date</b>",
                small_style,
            ),
            make_pdf_paragraph(
                report_date,
                normal_style,
            ),
        ],

        [
            Paragraph(
                "<b>Working Hours</b>",
                small_style,
            ),
            make_pdf_paragraph(
                f"{start_time} - {finish_time}",
                normal_style,
            ),
            Paragraph(
                "<b>Status</b>",
                small_style,
            ),
            make_pdf_paragraph(
                report_status.upper(),
                normal_style,
            ),
        ],
    ]

    info_table = Table(
        info_data,
        colWidths=[
            28 * mm,
            57 * mm,
            27 * mm,
            58 * mm,
        ],
    )

    info_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor(
                        "#F3F4F6"
                    ),
                ),

                (
                    "BACKGROUND",
                    (2, 0),
                    (2, -1),
                    colors.HexColor(
                        "#F3F4F6"
                    ),
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor(
                        "#D1D5DB"
                    ),
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
            ]
        )
    )

    story.append(
        info_table
    )


    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "Daily Summary",
            section_style,
        )
    )

    locations = set()

    for item in activities:

        location = (
            str(
                item.get(
                    "Location / Plot",
                    "",
                )
            )
            .strip()
        )

        if location:
            locations.add(location)

    summary_data = [
        [
            "MANPOWER",
            "PLANT",
            "ACTIVITIES",
            "LOCATIONS",
        ],

        [
            str(
                calculate_total_manpower(
                    report_key
                )
            ),

            str(
                calculate_total_equipment(
                    report_key
                )
            ),

            str(
                len(
                    activities
                )
            ),

            str(
                len(
                    locations
                )
            ),
        ],

        [
            "DELIVERIES",
            "INSPECTIONS",
            "ISSUES",
            "PHOTOS",
        ],

        [
            str(
                len(
                    deliveries
                )
            ),

            str(
                len(
                    inspections
                )
            ),

            str(
                len(
                    issues
                )
            ),

            str(
                len(
                    photos
                )
            ),
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            42.5 * mm,
            42.5 * mm,
            42.5 * mm,
            42.5 * mm,
        ],
    )

    summary_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#374151"
                    ),
                ),

                (
                    "BACKGROUND",
                    (0, 2),
                    (-1, 2),
                    colors.HexColor(
                        "#374151"
                    ),
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),

                (
                    "TEXTCOLOR",
                    (0, 2),
                    (-1, 2),
                    colors.white,
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),

                (
                    "FONTNAME",
                    (0, 2),
                    (-1, 2),
                    "Helvetica-Bold",
                ),

                (
                    "FONTNAME",
                    (0, 1),
                    (-1, 1),
                    "Helvetica-Bold",
                ),

                (
                    "FONTNAME",
                    (0, 3),
                    (-1, 3),
                    "Helvetica-Bold",
                ),

                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8,
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER",
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor(
                        "#D1D5DB"
                    ),
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5,
                ),
            ]
        )
    )

    story.append(
        summary_table
    )


    # --------------------------------------------------------
    # GENERIC TABLE FUNCTION
    # --------------------------------------------------------

    def add_section_table(
        title,
        headers,
        rows,
        widths,
    ):

        story.append(
            Paragraph(
                title,
                section_style,
            )
        )

        if not rows:

            story.append(
                Paragraph(
                    "No records.",
                    normal_style,
                )
            )

            return

        table_data = [
            [
                Paragraph(
                    f"<b>{escape(str(header))}</b>",
                    small_style,
                )
                for header
                in headers
            ]
        ]

        for row in rows:

            table_data.append(
                [
                    make_pdf_paragraph(
                        value,
                        small_style,
                    )
                    for value
                    in row
                ]
            )

        section_table = Table(
            table_data,
            colWidths=widths,
            repeatRows=1,
        )

        section_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor(
                            "#E5E7EB"
                        ),
                    ),

                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.35,
                        colors.HexColor(
                            "#D1D5DB"
                        ),
                    ),

                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "TOP",
                    ),

                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),

                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),

                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),

                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),
                ]
            )
        )

        story.append(
            section_table
        )


    # --------------------------------------------------------
    # MANPOWER
    # --------------------------------------------------------

    manpower_rows = []

    for item in manpower:

        manpower_rows.append(
            [
                item.get(
                    "Company",
                    "",
                ),

                item.get(
                    "Role",
                    "",
                ),

                item.get(
                    "Quantity",
                    0,
                ),
            ]
        )

    add_section_table(
        "Manpower",
        [
            "Company / Subcontractor",
            "Role / Trade",
            "Qty",
        ],
        manpower_rows,
        [
            65 * mm,
            80 * mm,
            25 * mm,
        ],
    )


    # --------------------------------------------------------
    # PLANT
    # --------------------------------------------------------

    equipment_rows = []

    for item in equipment:

        equipment_rows.append(
            [
                item.get(
                    "Equipment",
                    "",
                ),

                item.get(
                    "Company / Owner",
                    "",
                ),

                item.get(
                    "Quantity",
                    0,
                ),
            ]
        )

    add_section_table(
        "Plant & Equipment",
        [
            "Equipment",
            "Company / Owner",
            "Qty",
        ],
        equipment_rows,
        [
            75 * mm,
            70 * mm,
            25 * mm,
        ],
    )


    # --------------------------------------------------------
    # WEATHER
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "Weather & Workability",
            section_style,
        )
    )

    weather_data = [
        [
            Paragraph(
                "<b>Period</b>",
                small_style,
            ),

            Paragraph(
                "<b>Weather</b>",
                small_style,
            ),

            Paragraph(
                "<b>Workability</b>",
                small_style,
            ),
        ],

        [
            "Morning",
            report.get(
                "Weather Morning",
                "",
            ),
            report.get(
                "Workability Morning",
                "",
            ),
        ],

        [
            "Afternoon",
            report.get(
                "Weather Afternoon",
                "",
            ),
            report.get(
                "Workability Afternoon",
                "",
            ),
        ],
    ]

    weather_table = Table(
        weather_data,
        colWidths=[
            40 * mm,
            65 * mm,
            65 * mm,
        ],
    )

    weather_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#E5E7EB"
                    ),
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.35,
                    colors.HexColor(
                        "#D1D5DB"
                    ),
                ),

                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8,
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )
    )

    story.append(
        weather_table
    )

    weather_notes = report.get(
        "Weather Notes",
        "",
    )

    if weather_notes:

        story.append(
            Spacer(
                1,
                1.5 * mm,
            )
        )

        story.append(
            Paragraph(
                (
                    "<b>Weather Notes:</b> "
                    f"{pdf_text(weather_notes)}"
                ),
                normal_style,
            )
        )


    # --------------------------------------------------------
    # WORK PERFORMED
    # --------------------------------------------------------

    activity_rows = []

    for item in activities:

        activity_rows.append(
            [
                item.get(
                    "Location / Plot",
                    "",
                ),

                item.get(
                    "Company",
                    "",
                ),

                item.get(
                    "Activity",
                    "",
                ),

                item.get(
                    "Notes",
                    "",
                ),
            ]
        )

    add_section_table(
        "Work Performed",
        [
            "Location",
            "Company",
            "Activity",
            "Notes",
        ],
        activity_rows,
        [
            26 * mm,
            38 * mm,
            72 * mm,
            34 * mm,
        ],
    )


    # --------------------------------------------------------
    # DELIVERIES
    # --------------------------------------------------------

    delivery_rows = []

    for item in deliveries:

        delivery_rows.append(
            [
                item.get(
                    "Material / Delivery",
                    "",
                ),

                item.get(
                    "Quantity",
                    "",
                ),

                item.get(
                    "Supplier",
                    "",
                ),

                item.get(
                    "Location",
                    "",
                ),
            ]
        )

    add_section_table(
        "Deliveries / Materials",
        [
            "Material / Delivery",
            "Quantity",
            "Supplier",
            "Location",
        ],
        delivery_rows,
        [
            60 * mm,
            35 * mm,
            45 * mm,
            30 * mm,
        ],
    )


    # --------------------------------------------------------
    # INSPECTIONS
    # --------------------------------------------------------

    inspection_rows = []

    for item in inspections:

        inspection_rows.append(
            [
                item.get(
                    "Type",
                    "",
                ),

                item.get(
                    "Location",
                    "",
                ),

                item.get(
                    "Status",
                    "",
                ),

                item.get(
                    "Notes",
                    "",
                ),
            ]
        )

    add_section_table(
        "Inspections / Surveys",
        [
            "Type",
            "Location",
            "Status",
            "Notes",
        ],
        inspection_rows,
        [
            55 * mm,
            30 * mm,
            32 * mm,
            53 * mm,
        ],
    )


    # --------------------------------------------------------
    # ISSUES
    # --------------------------------------------------------

    issue_rows = []

    for item in issues:

        issue_rows.append(
            [
                item.get(
                    "Issue",
                    "",
                ),

                item.get(
                    "Location",
                    "",
                ),

                item.get(
                    "Impact",
                    "",
                ),

                item.get(
                    "Action / Follow-up",
                    "",
                ),
            ]
        )

    add_section_table(
        "Issues / Delays / Constraints",
        [
            "Issue",
            "Location",
            "Impact",
            "Action / Follow-up",
        ],
        issue_rows,
        [
            53 * mm,
            26 * mm,
            43 * mm,
            48 * mm,
        ],
    )


    # --------------------------------------------------------
    # GENERAL NOTES
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "General Notes",
            section_style,
        )
    )

    notes = report.get(
        "Notes",
        "",
    )

    if notes:

        story.append(
            Paragraph(
                pdf_text(
                    notes
                ),
                normal_style,
            )
        )

    else:

        story.append(
            Paragraph(
                "No additional notes.",
                normal_style,
            )
        )


    # --------------------------------------------------------
    # PHOTOS
    # --------------------------------------------------------

    if photos:

        story.append(
            PageBreak()
        )

        story.append(
            Paragraph(
                "Site Photographs",
                title_style,
            )
        )

        story.append(
            Paragraph(
                (
                    f"{pdf_text(project)} | "
                    f"{pdf_text(report_date)}"
                ),
                subtitle_style,
            )
        )

        photo_cells = []

        for photo in photos:

            try:

                image_bytes = photo.get(
                    "Bytes"
                )

                image_reader = ImageReader(
                    BytesIO(
                        image_bytes
                    )
                )

                image_width, image_height = (
                    image_reader.getSize()
                )

                max_width = (
                    79 * mm
                )

                max_height = (
                    74 * mm
                )

                ratio = min(
                    max_width
                    / image_width,

                    max_height
                    / image_height,
                )

                display_width = (
                    image_width
                    * ratio
                )

                display_height = (
                    image_height
                    * ratio
                )

                image = RLImage(
                    BytesIO(
                        image_bytes
                    ),
                    width=display_width,
                    height=display_height,
                )

                location = photo.get(
                    "Location",
                    "",
                )

                description = photo.get(
                    "Description",
                    "",
                )

                caption_parts = []

                if location:

                    caption_parts.append(
                        f"<b>{escape(str(location))}</b>"
                    )

                if description:

                    caption_parts.append(
                        escape(
                            str(description)
                        )
                    )

                if not caption_parts:

                    caption_parts.append(
                        escape(
                            str(
                                photo.get(
                                    "File Name",
                                    "Site Photo",
                                )
                            )
                        )
                    )

                photo_block = [
                    image,

                    Spacer(
                        1,
                        1.5 * mm,
                    ),

                    Paragraph(
                        "<br/>".join(
                            caption_parts
                        ),
                        photo_caption_style,
                    ),
                ]

                photo_table = Table(
                    [
                        [
                            photo_block
                        ]
                    ],
                    colWidths=[
                        82 * mm
                    ],
                )

                photo_table.setStyle(
                    TableStyle(
                        [
                            (
                                "BOX",
                                (0, 0),
                                (-1, -1),
                                0.5,
                                colors.HexColor(
                                    "#D1D5DB"
                                ),
                            ),

                            (
                                "BACKGROUND",
                                (0, 0),
                                (-1, -1),
                                colors.HexColor(
                                    "#F9FAFB"
                                ),
                            ),

                            (
                                "ALIGN",
                                (0, 0),
                                (-1, -1),
                                "CENTER",
                            ),

                            (
                                "VALIGN",
                                (0, 0),
                                (-1, -1),
                                "TOP",
                            ),

                            (
                                "TOPPADDING",
                                (0, 0),
                                (-1, -1),
                                6,
                            ),

                            (
                                "BOTTOMPADDING",
                                (0, 0),
                                (-1, -1),
                                6,
                            ),
                        ]
                    )
                )

                photo_cells.append(
                    photo_table
                )

            except Exception:

                continue

        if photo_cells:

            photo_rows = []

            for index in range(
                0,
                len(photo_cells),
                2,
            ):

                row = [
                    photo_cells[
                        index
                    ]
                ]

                if (
                    index + 1
                    < len(photo_cells)
                ):

                    row.append(
                        photo_cells[
                            index + 1
                        ]
                    )

                else:

                    row.append(
                        ""
                    )

                photo_rows.append(
                    row
                )

            photos_table = Table(
                photo_rows,
                colWidths=[
                    85 * mm,
                    85 * mm,
                ],
                hAlign="CENTER",
            )

            photos_table.setStyle(
                TableStyle(
                    [
                        (
                            "VALIGN",
                            (0, 0),
                            (-1, -1),
                            "TOP",
                        ),

                        (
                            "LEFTPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),

                        (
                            "RIGHTPADDING",
                            (0, 0),
                            (-1, -1),
                            2,
                        ),

                        (
                            "TOPPADDING",
                            (0, 0),
                            (-1, -1),
                            3,
                        ),

                        (
                            "BOTTOMPADDING",
                            (0, 0),
                            (-1, -1),
                            3,
                        ),
                    ]
                )
            )

            story.append(
                photos_table
            )


    # --------------------------------------------------------
    # BUILD PDF
    # --------------------------------------------------------

    doc.build(
        story,
        onFirstPage=add_pdf_page_number,
        onLaterPages=add_pdf_page_number,
    )

    output.seek(0)

    return output.getvalue()


# ============================================================
# EXCEL GENERATOR
# ============================================================

def generate_daily_excel(
    report_key,
):

    report = (
        st.session_state.daily_reports.get(
            report_key,
            {},
        )
    )

    manpower = daily_items_for_key(
        st.session_state.daily_manpower,
        report_key,
    )

    equipment = daily_items_for_key(
        st.session_state.daily_equipment,
        report_key,
    )

    activities = daily_items_for_key(
        st.session_state.daily_activities,
        report_key,
    )

    deliveries = daily_items_for_key(
        st.session_state.daily_deliveries,
        report_key,
    )

    inspections = daily_items_for_key(
        st.session_state.daily_inspections,
        report_key,
    )

    issues = daily_items_for_key(
        st.session_state.daily_issues,
        report_key,
    )

    photos = daily_items_for_key(
        st.session_state.daily_photos,
        report_key,
    )

    workbook = Workbook()

    summary = workbook.active

    summary.title = (
        "Daily Summary"
    )

    dark_fill = PatternFill(
        "solid",
        fgColor="374151",
    )

    orange_fill = PatternFill(
        "solid",
        fgColor="D15C26",
    )

    light_fill = PatternFill(
        "solid",
        fgColor="F3F4F6",
    )

    white_font = Font(
        color="FFFFFF",
        bold=True,
    )

    header_font = Font(
        bold=True,
        color="1F2937",
    )

    company_font = Font(
        bold=True,
        size=16,
        color="D15C26",
    )

    title_font = Font(
        bold=True,
        size=18,
        color="1F2937",
    )

    thin_side = Side(
        style="thin",
        color="D1D5DB",
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )


    # --------------------------------------------------------
    # SUMMARY SHEET
    # --------------------------------------------------------

    summary.merge_cells(
        "A1:B1"
    )

    summary[
        "A1"
    ] = "GLENVEAGH"

    summary[
        "A1"
    ].font = company_font

    summary.merge_cells(
        "C1:F1"
    )

    summary[
        "C1"
    ] = "SITE DAILY REPORT"

    summary[
        "C1"
    ].font = title_font

    summary[
        "C1"
    ].alignment = Alignment(
        horizontal="right"
    )

    summary[
        "A3"
    ] = "Project / Site"

    summary[
        "B3"
    ] = report.get(
        "Project",
        "",
    )

    summary[
        "D3"
    ] = "Date"

    summary[
        "E3"
    ] = format_date_for_display(
        report.get(
            "Date",
            "",
        )
    )

    summary[
        "A4"
    ] = "Working Hours"

    summary[
        "B4"
    ] = (
        f"{format_time_for_display(report.get('Working Start', ''))}"
        f" - "
        f"{format_time_for_display(report.get('Working Finish', ''))}"
    )

    summary[
        "D4"
    ] = "Status"

    summary[
        "E4"
    ] = report.get(
        "Status",
        "",
    )

    for cell in [
        "A3",
        "D3",
        "A4",
        "D4",
    ]:

        summary[
            cell
        ].font = header_font

        summary[
            cell
        ].fill = light_fill


    summary[
        "A6"
    ] = "DAILY SUMMARY"

    summary[
        "A6"
    ].fill = dark_fill

    summary[
        "A6"
    ].font = white_font

    summary.merge_cells(
        "A6:F6"
    )

    locations = set()

    for item in activities:

        location = (
            str(
                item.get(
                    "Location / Plot",
                    "",
                )
            )
            .strip()
        )

        if location:
            locations.add(location)

    summary_metrics = [
        (
            "Total Manpower",
            calculate_total_manpower(
                report_key
            ),
        ),

        (
            "Plant / Equipment",
            calculate_total_equipment(
                report_key
            ),
        ),

        (
            "Activities",
            len(
                activities
            ),
        ),

        (
            "Active Locations",
            len(
                locations
            ),
        ),

        (
            "Deliveries",
            len(
                deliveries
            ),
        ),

        (
            "Inspections",
            len(
                inspections
            ),
        ),

        (
            "Issues",
            len(
                issues
            ),
        ),

        (
            "Photos",
            len(
                photos
            ),
        ),
    ]

    row = 7

    for label, value in summary_metrics:

        summary.cell(
            row=row,
            column=1,
            value=label,
        )

        summary.cell(
            row=row,
            column=2,
            value=value,
        )

        summary.cell(
            row=row,
            column=1,
        ).font = header_font

        row += 1


    summary[
        "D7"
    ] = "Morning Weather"

    summary[
        "E7"
    ] = report.get(
        "Weather Morning",
        "",
    )

    summary[
        "D8"
    ] = "Morning Workability"

    summary[
        "E8"
    ] = report.get(
        "Workability Morning",
        "",
    )

    summary[
        "D9"
    ] = "Afternoon Weather"

    summary[
        "E9"
    ] = report.get(
        "Weather Afternoon",
        "",
    )

    summary[
        "D10"
    ] = "Afternoon Workability"

    summary[
        "E10"
    ] = report.get(
        "Workability Afternoon",
        "",
    )

    summary[
        "D11"
    ] = "Weather Notes"

    summary[
        "E11"
    ] = report.get(
        "Weather Notes",
        "",
    )

    for cell in [
        "D7",
        "D8",
        "D9",
        "D10",
        "D11",
    ]:

        summary[
            cell
        ].font = header_font


    summary[
        "A17"
    ] = "GENERAL NOTES"

    summary[
        "A17"
    ].fill = dark_fill

    summary[
        "A17"
    ].font = white_font

    summary.merge_cells(
        "A17:F17"
    )

    summary.merge_cells(
        "A18:F22"
    )

    summary[
        "A18"
    ] = report.get(
        "Notes",
        "",
    )

    summary[
        "A18"
    ].alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )


    for column, width in {
        "A": 24,
        "B": 25,
        "C": 8,
        "D": 24,
        "E": 28,
        "F": 15,
    }.items():

        summary.column_dimensions[
            column
        ].width = width


    # --------------------------------------------------------
    # GENERIC EXCEL SHEET
    # --------------------------------------------------------

    def create_data_sheet(
        sheet_name,
        columns,
        rows,
    ):

        ws = workbook.create_sheet(
            title=sheet_name
        )

        ws.freeze_panes = (
            "A2"
        )

        for col_index, header in enumerate(
            columns,
            start=1,
        ):

            cell = ws.cell(
                row=1,
                column=col_index,
                value=header,
            )

            cell.fill = dark_fill
            cell.font = white_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            cell.border = border

        for row_index, row_data in enumerate(
            rows,
            start=2,
        ):

            for col_index, value in enumerate(
                row_data,
                start=1,
            ):

                cell = ws.cell(
                    row=row_index,
                    column=col_index,
                    value=value,
                )

                cell.border = border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for col_index, header in enumerate(
            columns,
            start=1,
        ):

            max_length = len(
                str(header)
            )

            for row_index in range(
                2,
                ws.max_row + 1,
            ):

                value = ws.cell(
                    row=row_index,
                    column=col_index,
                ).value

                if value is not None:

                    max_length = max(
                        max_length,
                        min(
                            len(
                                str(value)
                            ),
                            60,
                        ),
                    )

            ws.column_dimensions[
                get_column_letter(
                    col_index
                )
            ].width = (
                max_length
                + 3
            )

        return ws


    # --------------------------------------------------------
    # MANPOWER
    # --------------------------------------------------------

    create_data_sheet(
        "Manpower",
        [
            "Company / Subcontractor",
            "Role / Trade",
            "Quantity",
        ],
        [
            [
                item.get(
                    "Company",
                    "",
                ),

                item.get(
                    "Role",
                    "",
                ),

                item.get(
                    "Quantity",
                    0,
                ),
            ]
            for item
            in manpower
        ],
    )


    # --------------------------------------------------------
    # PLANT
    # --------------------------------------------------------

    create_data_sheet(
        "Plant & Equipment",
        [
            "Equipment",
            "Company / Owner",
            "Quantity",
        ],
        [
            [
                item.get(
                    "Equipment",
                    "",
                ),

                item.get(
                    "Company / Owner",
                    "",
                ),

                item.get(
                    "Quantity",
                    0,
                ),
            ]
            for item
            in equipment
        ],
    )


    # --------------------------------------------------------
    # WORK PERFORMED
    # --------------------------------------------------------

    create_data_sheet(
        "Work Performed",
        [
            "Location / Plot",
            "Company",
            "Activity",
            "Notes",
        ],
        [
            [
                item.get(
                    "Location / Plot",
                    "",
                ),

                item.get(
                    "Company",
                    "",
                ),

                item.get(
                    "Activity",
                    "",
                ),

                item.get(
                    "Notes",
                    "",
                ),
            ]
            for item
            in activities
        ],
    )


    # --------------------------------------------------------
    # DELIVERIES
    # --------------------------------------------------------

    create_data_sheet(
        "Deliveries",
        [
            "Material / Delivery",
            "Quantity",
            "Supplier",
            "Location",
        ],
        [
            [
                item.get(
                    "Material / Delivery",
                    "",
                ),

                item.get(
                    "Quantity",
                    "",
                ),

                item.get(
                    "Supplier",
                    "",
                ),

                item.get(
                    "Location",
                    "",
                ),
            ]
            for item
            in deliveries
        ],
    )


    # --------------------------------------------------------
    # INSPECTIONS
    # --------------------------------------------------------

    create_data_sheet(
        "Inspections",
        [
            "Type",
            "Location",
            "Status",
            "Notes",
        ],
        [
            [
                item.get(
                    "Type",
                    "",
                ),

                item.get(
                    "Location",
                    "",
                ),

                item.get(
                    "Status",
                    "",
                ),

                item.get(
                    "Notes",
                    "",
                ),
            ]
            for item
            in inspections
        ],
    )


    # --------------------------------------------------------
    # ISSUES
    # --------------------------------------------------------

    create_data_sheet(
        "Issues",
        [
            "Issue",
            "Location",
            "Impact",
            "Action / Follow-up",
        ],
        [
            [
                item.get(
                    "Issue",
                    "",
                ),

                item.get(
                    "Location",
                    "",
                ),

                item.get(
                    "Impact",
                    "",
                ),

                item.get(
                    "Action / Follow-up",
                    "",
                ),
            ]
            for item
            in issues
        ],
    )


    # --------------------------------------------------------
    # PHOTOS
    # --------------------------------------------------------

    create_data_sheet(
        "Photos",
        [
            "File Name",
            "Location",
            "Description",
        ],
        [
            [
                item.get(
                    "File Name",
                    "",
                ),

                item.get(
                    "Location",
                    "",
                ),

                item.get(
                    "Description",
                    "",
                ),
            ]
            for item
            in photos
        ],
    )


    # --------------------------------------------------------
    # SAVE
    # --------------------------------------------------------

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output.getvalue()


# ============================================================
# HEADER
# ============================================================

st.title(
    "🏗️ Site Engineer Tool"
)

st.caption(
    "Survey, QA/QC, inspections, checklists, "
    "as-built records and site management."
)


# ============================================================
# MENU
# ============================================================

st.sidebar.title(
    "Site Engineer Tool"
)

pagina = st.sidebar.radio(
    "Menu",
    [
        "Dashboard",
        "Surveys",
        "Inspections",
        "Checklists",
        "Snag List",
        "As-Built",
        "Levels",
        "Concrete",
        "Drainage",
        "Public Lighting",
        "Daily Report",
    ],
)


# ============================================================
# DASHBOARD
# ============================================================

if pagina == "Dashboard":

    st.header(
        "Dashboard"
    )

    total_surveys = (
        len(
            st.session_state.top_block_points
        )
        + len(
            st.session_state.dimension_checks
        )
        + len(
            st.session_state.floor_slab_points
        )
    )

    total_no_pass = 0

    for registro in (
        st.session_state.top_block_points
    ):

        if (
            registro.get(
                "Status"
            )
            == "NO PASS"
        ):

            total_no_pass += 1


    for registro in (
        st.session_state.dimension_checks
    ):

        if (
            registro.get(
                "Status"
            )
            == "NO PASS"
        ):

            total_no_pass += 1


    for registro in (
        st.session_state.floor_slab_points
    ):

        if (
            registro.get(
                "Status"
            )
            == "NO PASS"
        ):

            total_no_pass += 1


    c1, c2, c3, c4 = (
        st.columns(4)
    )

    with c1:

        st.metric(
            "Open Items",
            0,
        )

    with c2:

        st.metric(
            "Pending Inspections",
            0,
        )

    with c3:

        st.metric(
            "Snags",
            0,
        )

    with c4:

        st.metric(
            "Survey Records",
            total_surveys,
        )


    st.divider()

    st.subheader(
        "Project Overview"
    )

    d1, d2, d3, d4 = (
        st.columns(4)
    )

    with d1:

        st.metric(
            "Survey Checks",
            total_surveys,
        )

    with d2:

        st.metric(
            "Items Requiring Review",
            total_no_pass,
        )

    with d3:

        st.metric(
            "Foundation Checklists",
            len(
                st.session_state.foundation_checklists
            ),
        )

    with d4:

        st.metric(
            "Daily Reports",
            len(
                st.session_state.daily_reports
            ),
        )


# ============================================================
# SURVEYS
# ============================================================

elif pagina == "Surveys":

    st.header(
        "Surveys"
    )

    st.caption(
        "House dimensional, level and construction surveys."
    )

    survey_type = st.selectbox(
        "Survey Type",
        [
            "Top of Block / Level",
            "Dimensions & Squareness",
            "Floor Slab Level",
        ],
    )

    st.divider()


    # ========================================================
    # TOP OF BLOCK
    # ========================================================

    if (
        survey_type
        == "Top of Block / Level"
    ):

        st.subheader(
            "Top of Block / Level Survey"
        )

        st.caption(
            "Absolute GPS elevation and relative "
            "laser perimeter level check."
        )

        c1, c2, c3 = (
            st.columns(3)
        )

        with c1:

            tob_date = (
                st.date_input(
                    "Date",
                    value=date.today(),
                    key="tob_date",
                )
            )

        with c2:

            tob_project = (
                st.text_input(
                    "Project / Site",
                    key="tob_project",
                    placeholder="Ex.: M1.02",
                )
            )

        with c3:

            tob_house = (
                st.text_input(
                    "Plot / House",
                    key="tob_house",
                    placeholder="Ex.: BT05",
                )
            )


        c4, c5, c6 = (
            st.columns(3)
        )

        with c4:

            design_level = (
                st.number_input(
                    "Design Reference Level (m)",
                    value=0.000,
                    step=0.001,
                    format="%.3f",
                    key="tob_design",
                )
            )

        with c5:

            gps_level = (
                st.number_input(
                    "GPS Top of Block Level (m)",
                    value=0.000,
                    step=0.001,
                    format="%.3f",
                    key="tob_gps",
                )
            )

        with c6:

            gps_tolerance = (
                st.number_input(
                    "GPS Tolerance (± mm)",
                    min_value=0.0,
                    value=8.0,
                    step=1.0,
                    key="tob_gps_tol",
                )
            )


        gps_difference = (
            difference_mm(
                gps_level,
                design_level,
            )
        )

        gps_status = (
            check_tolerance(
                gps_difference,
                gps_tolerance,
            )
        )


        g1, g2 = (
            st.columns(2)
        )

        with g1:

            st.metric(
                "GPS Difference",
                f"{gps_difference:+.1f} mm",
            )

        with g2:

            st.metric(
                "GPS Status",
                status_icon(
                    gps_status
                ),
            )


        st.divider()

        st.markdown(
            "### Laser Perimeter Level Check"
        )

        st.caption(
            "Adicione quantos pontos forem necessários. "
            "Não existe quantidade fixa de pontos."
        )


        with st.form(
            "add_top_block_point",
            clear_on_submit=True,
        ):

            p1, p2, p3 = (
                st.columns(3)
            )

            with p1:

                point = (
                    st.text_input(
                        "Point / Wall Reference",
                        placeholder=(
                            "Ex.: Wall 01 / P01"
                        ),
                    )
                )

            with p2:

                laser_difference = (
                    st.number_input(
                        "Laser Difference (mm)",
                        value=0.0,
                        step=1.0,
                    )
                )

            with p3:

                laser_tolerance = (
                    st.number_input(
                        "Tolerance (± mm)",
                        min_value=0.0,
                        value=8.0,
                        step=1.0,
                    )
                )

            notes = (
                st.text_input(
                    "Notes",
                    placeholder="Optional",
                )
            )

            add_point = (
                st.form_submit_button(
                    "Add Level Point"
                )
            )


        if add_point:

            status = (
                check_tolerance(
                    laser_difference,
                    laser_tolerance,
                )
            )

            st.session_state.top_block_points.append(
                {
                    "Date":
                        tob_date,

                    "Project":
                        tob_project,

                    "Plot / House":
                        tob_house,

                    "Reference Level (m)":
                        design_level,

                    "GPS Level (m)":
                        gps_level,

                    "GPS Difference (mm)":
                        round(
                            gps_difference,
                            1,
                        ),

                    "Point":
                        point,

                    "Laser Difference (mm)":
                        round(
                            laser_difference,
                            1,
                        ),

                    "Tolerance (mm)":
                        laser_tolerance,

                    "Status":
                        status,

                    "Notes":
                        notes,
                }
            )

            st.rerun()


        current_point_indices = [
            index
            for index, row
            in enumerate(
                st.session_state.top_block_points
            )
            if (
                not tob_house
                or
                row.get(
                    "Plot / House"
                )
                == tob_house
            )
        ]

        current_points = [
            st.session_state.top_block_points[
                index
            ]
            for index
            in current_point_indices
        ]


        if current_points:

            df_top = (
                pd.DataFrame(
                    current_points
                )
            )

            df_top_display = (
                df_top.copy()
            )

            df_top_display[
                "Status"
            ] = (
                df_top_display[
                    "Status"
                ].apply(
                    lambda value:
                    "✅ PASS"
                    if value == "PASS"
                    else "⚠️ NO PASS"
                )
            )


            st.subheader(
                "Recorded Perimeter Points"
            )

            st.dataframe(
                df_top_display,
                use_container_width=True,
                hide_index=True,
            )


            with st.expander(
                "Edit / Delete recorded perimeter points"
            ):

                for (
                    display_number,
                    original_index,
                ) in enumerate(
                    current_point_indices,
                    start=1,
                ):

                    row = (
                        st.session_state.top_block_points[
                            original_index
                        ]
                    )

                    d1, d2, d3, d4 = (
                        st.columns(
                            [
                                1,
                                3,
                                2,
                                0.7,
                            ]
                        )
                    )

                    with d1:

                        st.write(
                            display_number
                        )

                    with d2:

                        st.write(
                            row.get(
                                "Point",
                                "Point",
                            )
                        )

                    with d3:

                        st.write(
                            status_icon(
                                row.get(
                                    "Status"
                                )
                            )
                        )

                    with d4:

                        if st.button(
                            "🗑️",
                            key=(
                                f"delete_tob_"
                                f"{original_index}"
                            ),
                        ):

                            del (
                                st.session_state.top_block_points[
                                    original_index
                                ]
                            )

                            st.rerun()


            laser_values = (
                df_top[
                    "Laser Difference (mm)"
                ]
            )

            highest = (
                laser_values.max()
            )

            lowest = (
                laser_values.min()
            )


            r1, r2, r3 = (
                st.columns(3)
            )

            with r1:

                st.metric(
                    "Points Checked",
                    len(
                        df_top
                    ),
                )

            with r2:

                st.metric(
                    "Highest",
                    f"{highest:+.1f} mm",
                )

            with r3:

                st.metric(
                    "Lowest",
                    f"{lowest:+.1f} mm",
                )


            failed = (
                df_top[
                    df_top[
                        "Status"
                    ]
                    == "NO PASS"
                ]
            )


            if not failed.empty:

                st.markdown(
                    "### ⚠️ Points Requiring Review"
                )

                failed_display = (
                    failed[
                        [
                            "Point",
                            "Laser Difference (mm)",
                            "Tolerance (mm)",
                        ]
                    ]
                    .copy()
                )

                failed_display[
                    "Status"
                ] = "⚠️ NO PASS"

                st.dataframe(
                    failed_display,
                    use_container_width=True,
                    hide_index=True,
                )


            if (
                failed.empty
                and
                gps_status == "PASS"
            ):

                st.success(
                    "✅ Top of Block Survey: PASS"
                )

            else:

                st.warning(
                    "⚠️ Top of Block Survey: "
                    "REVIEW REQUIRED"
                )


    # ========================================================
    # DIMENSIONS
    # ========================================================

    elif (
        survey_type
        == "Dimensions & Squareness"
    ):

        st.subheader(
            "Dimensions & Squareness Survey"
        )

        c1, c2, c3 = (
            st.columns(3)
        )

        with c1:

            dim_date = (
                st.date_input(
                    "Date",
                    value=date.today(),
                    key="dim_date",
                )
            )

        with c2:

            dim_project = (
                st.text_input(
                    "Project / Site",
                    key="dim_project",
                )
            )

        with c3:

            dim_house = (
                st.text_input(
                    "Plot / House",
                    key="dim_house",
                )
            )


        with st.form(
            "dimension_form",
            clear_on_submit=True,
        ):

            d1, d2 = (
                st.columns(2)
            )

            with d1:

                reference = (
                    st.text_input(
                        "Reference"
                    )
                )

            with d2:

                room = (
                    st.text_input(
                        "Room / Area"
                    )
                )


            d3, d4, d5 = (
                st.columns(3)
            )

            with d3:

                design_dimension = (
                    st.number_input(
                        "Design Dimension (m)",
                        min_value=0.0,
                        value=0.000,
                        step=0.001,
                        format="%.3f",
                    )
                )

            with d4:

                measured_dimension = (
                    st.number_input(
                        "Measured Dimension (m)",
                        min_value=0.0,
                        value=0.000,
                        step=0.001,
                        format="%.3f",
                    )
                )

            with d5:

                dim_tolerance = (
                    st.number_input(
                        "Tolerance (± mm)",
                        min_value=0.0,
                        value=8.0,
                        step=1.0,
                    )
                )


            add_dimension = (
                st.form_submit_button(
                    "Add Dimension"
                )
            )


        if add_dimension:

            diff = (
                difference_mm(
                    measured_dimension,
                    design_dimension,
                )
            )

            status = (
                check_tolerance(
                    diff,
                    dim_tolerance,
                )
            )

            st.session_state.dimension_checks.append(
                {
                    "Date":
                        dim_date,

                    "Project":
                        dim_project,

                    "Plot / House":
                        dim_house,

                    "Room / Area":
                        room,

                    "Reference":
                        reference,

                    "Design (m)":
                        design_dimension,

                    "Measured (m)":
                        measured_dimension,

                    "Difference (mm)":
                        round(
                            diff,
                            1,
                        ),

                    "Tolerance (mm)":
                        dim_tolerance,

                    "Status":
                        status,
                }
            )

            st.rerun()


        current_dimension_indices = [
            index
            for index, row
            in enumerate(
                st.session_state.dimension_checks
            )
            if (
                not dim_house
                or
                row.get(
                    "Plot / House"
                )
                == dim_house
            )
        ]


        current_dimensions = [
            st.session_state.dimension_checks[
                index
            ]
            for index
            in current_dimension_indices
        ]


        if current_dimensions:

            df_dimensions = (
                pd.DataFrame(
                    current_dimensions
                )
            )

            display_dimensions = (
                df_dimensions.copy()
            )

            display_dimensions[
                "Status"
            ] = (
                display_dimensions[
                    "Status"
                ].apply(
                    lambda value:
                    "✅ PASS"
                    if value == "PASS"
                    else "⚠️ NO PASS"
                )
            )


            st.subheader(
                "Recorded Dimensions"
            )

            st.dataframe(
                display_dimensions,
                use_container_width=True,
                hide_index=True,
            )


            with st.expander(
                "Edit / Delete recorded dimensions"
            ):

                for (
                    number,
                    original_index,
                ) in enumerate(
                    current_dimension_indices,
                    start=1,
                ):

                    row = (
                        st.session_state.dimension_checks[
                            original_index
                        ]
                    )

                    x1, x2, x3, x4 = (
                        st.columns(
                            [
                                1,
                                3,
                                2,
                                0.7,
                            ]
                        )
                    )

                    with x1:

                        st.write(
                            number
                        )

                    with x2:

                        st.write(
                            row.get(
                                "Reference",
                                "Dimension",
                            )
                        )

                    with x3:

                        st.write(
                            status_icon(
                                row.get(
                                    "Status"
                                )
                            )
                        )

                    with x4:

                        if st.button(
                            "🗑️",
                            key=(
                                f"delete_dimension_"
                                f"{original_index}"
                            ),
                        ):

                            del (
                                st.session_state.dimension_checks[
                                    original_index
                                ]
                            )

                            st.rerun()


        st.divider()

        st.markdown(
            "### Squareness / Diagonal Check"
        )


        s1, s2 = (
            st.columns(2)
        )

        with s1:

            design_length = (
                st.number_input(
                    "Design Length (m)",
                    min_value=0.0,
                    value=0.000,
                    step=0.001,
                    format="%.3f",
                )
            )

        with s2:

            design_width = (
                st.number_input(
                    "Design Width (m)",
                    min_value=0.0,
                    value=0.000,
                    step=0.001,
                    format="%.3f",
                )
            )


        s3, s4, s5 = (
            st.columns(3)
        )

        with s3:

            diagonal_a = (
                st.number_input(
                    "Measured Diagonal A (m)",
                    min_value=0.0,
                    value=0.000,
                    step=0.001,
                    format="%.3f",
                )
            )

        with s4:

            diagonal_b = (
                st.number_input(
                    "Measured Diagonal B (m)",
                    min_value=0.0,
                    value=0.000,
                    step=0.001,
                    format="%.3f",
                )
            )

        with s5:

            diagonal_tolerance = (
                st.number_input(
                    "Diagonal Tolerance (mm)",
                    min_value=0.0,
                    value=8.0,
                    step=1.0,
                )
            )


        theoretical = (
            calculate_diagonal(
                design_length,
                design_width,
            )
        )

        diagonal_difference = (
            abs(
                diagonal_a
                - diagonal_b
            )
            * 1000
        )

        square_status = (
            "PASS"
            if (
                diagonal_difference
                <= diagonal_tolerance
            )
            else
            "NO PASS"
        )


        q1, q2, q3 = (
            st.columns(3)
        )

        with q1:

            st.metric(
                "Theoretical Diagonal",
                (
                    f"{theoretical:.3f} m"
                    if theoretical
                    else "-"
                ),
            )

        with q2:

            st.metric(
                "Diagonal Difference",
                f"{diagonal_difference:.1f} mm",
            )

        with q3:

            st.metric(
                "Squareness",
                status_icon(
                    square_status
                ),
            )


    # ========================================================
    # FLOOR SLAB
    # ========================================================

    elif (
        survey_type
        == "Floor Slab Level"
    ):

        st.subheader(
            "Floor Slab Level Survey"
        )


        c1, c2, c3 = (
            st.columns(3)
        )

        with c1:

            slab_date = (
                st.date_input(
                    "Date",
                    value=date.today(),
                    key="slab_date",
                )
            )

        with c2:

            slab_project = (
                st.text_input(
                    "Project / Site",
                    key="slab_project",
                )
            )

        with c3:

            slab_house = (
                st.text_input(
                    "Plot / House",
                    key="slab_house",
                )
            )


        c4, c5 = (
            st.columns(2)
        )

        with c4:

            slab_reference = (
                st.number_input(
                    "Reference Level / Datum (m)",
                    value=0.000,
                    step=0.001,
                    format="%.3f",
                    key="slab_reference",
                )
            )

        with c5:

            slab_tolerance = (
                st.number_input(
                    "Tolerance (± mm)",
                    min_value=0.0,
                    value=8.0,
                    step=1.0,
                    key="slab_tolerance",
                )
            )


        with st.form(
            "add_slab_point",
            clear_on_submit=True,
        ):

            f1, f2 = (
                st.columns(2)
            )

            with f1:

                slab_point = (
                    st.text_input(
                        "Point",
                        placeholder="Ex.: F01",
                    )
                )

            with f2:

                slab_difference = (
                    st.number_input(
                        "Laser Difference (mm)",
                        value=0.0,
                        step=1.0,
                    )
                )

            slab_notes = (
                st.text_input(
                    "Notes",
                    placeholder="Optional",
                )
            )

            add_slab = (
                st.form_submit_button(
                    "Add Slab Point"
                )
            )


        if add_slab:

            status = (
                check_tolerance(
                    slab_difference,
                    slab_tolerance,
                )
            )

            st.session_state.floor_slab_points.append(
                {
                    "Date":
                        slab_date,

                    "Project":
                        slab_project,

                    "Plot / House":
                        slab_house,

                    "Reference Level (m)":
                        slab_reference,

                    "Point":
                        slab_point,

                    "Difference (mm)":
                        round(
                            slab_difference,
                            1,
                        ),

                    "Tolerance (mm)":
                        slab_tolerance,

                    "Status":
                        status,

                    "Notes":
                        slab_notes,
                }
            )

            st.rerun()


        current_slab_indices = [
            index
            for index, row
            in enumerate(
                st.session_state.floor_slab_points
            )
            if (
                not slab_house
                or
                row.get(
                    "Plot / House"
                )
                == slab_house
            )
        ]


        current_slab = [
            st.session_state.floor_slab_points[
                index
            ]
            for index
            in current_slab_indices
        ]


        if current_slab:

            df_slab = (
                pd.DataFrame(
                    current_slab
                )
            )

            display_slab = (
                df_slab.copy()
            )

            display_slab[
                "Status"
            ] = (
                display_slab[
                    "Status"
                ].apply(
                    lambda value:
                    "✅ PASS"
                    if value == "PASS"
                    else "⚠️ NO PASS"
                )
            )


            st.subheader(
                "Recorded Slab Points"
            )

            st.dataframe(
                display_slab,
                use_container_width=True,
                hide_index=True,
            )


            with st.expander(
                "Edit / Delete recorded slab points"
            ):

                for (
                    number,
                    original_index,
                ) in enumerate(
                    current_slab_indices,
                    start=1,
                ):

                    row = (
                        st.session_state.floor_slab_points[
                            original_index
                        ]
                    )

                    x1, x2, x3, x4 = (
                        st.columns(
                            [
                                1,
                                3,
                                2,
                                0.7,
                            ]
                        )
                    )

                    with x1:

                        st.write(
                            number
                        )

                    with x2:

                        st.write(
                            row.get(
                                "Point",
                                "Point",
                            )
                        )

                    with x3:

                        st.write(
                            status_icon(
                                row.get(
                                    "Status"
                                )
                            )
                        )

                    with x4:

                        if st.button(
                            "🗑️",
                            key=(
                                f"delete_slab_"
                                f"{original_index}"
                            ),
                        ):

                            del (
                                st.session_state.floor_slab_points[
                                    original_index
                                ]
                            )

                            st.rerun()


            values = (
                df_slab[
                    "Difference (mm)"
                ]
            )


            z1, z2, z3 = (
                st.columns(3)
            )

            with z1:

                st.metric(
                    "Points Checked",
                    len(
                        df_slab
                    ),
                )

            with z2:

                st.metric(
                    "Highest",
                    f"{values.max():+.1f} mm",
                )

            with z3:

                st.metric(
                    "Lowest",
                    f"{values.min():+.1f} mm",
                )


            failed_slab = (
                df_slab[
                    df_slab[
                        "Status"
                    ]
                    == "NO PASS"
                ]
            )

            if not failed_slab.empty:

                st.warning(
                    "⚠️ Floor Slab Survey: "
                    "REVIEW REQUIRED"
                )


# ============================================================
# INSPECTIONS
# ============================================================

elif pagina == "Inspections":

    st.header(
        "Inspections"
    )

    st.info(
        "Módulo Inspections em construção."
    )


# ============================================================
# CHECKLISTS
# ============================================================

elif pagina == "Checklists":

    st.header(
        "Checklists"
    )

    st.caption(
        "Daily construction and QA/QC checklists."
    )

    checklist_type = (
        st.selectbox(
            "Checklist Type",
            [
                "Foundation",
            ],
        )
    )

    st.divider()


    if (
        checklist_type
        == "Foundation"
    ):

        st.header(
            "Foundation Checklist"
        )

        st.caption(
            "Create a new foundation record or reopen "
            "an existing foundation to continue updating it."
        )


        foundation_mode = (
            st.radio(
                "Foundation Record",
                [
                    "➕ New Foundation Checklist",
                    "📂 Open Existing Checklist",
                ],
                horizontal=True,
            )
        )


        existing_keys = list(
            st.session_state
            .foundation_checklists
            .keys()
        )


        # ====================================================
        # OPEN FOUNDATION
        # ====================================================

        if (
            foundation_mode
            == "📂 Open Existing Checklist"
        ):

            if not existing_keys:

                st.info(
                    "No Foundation Checklists "
                    "have been created yet."
                )

                st.stop()


            summary_rows = []

            for key in existing_keys:

                record = (
                    st.session_state
                    .foundation_checklists[
                        key
                    ]
                )


                bottom_points_summary = [
                    point
                    for point
                    in st.session_state
                    .foundation_bottom_points
                    if (
                        point.get(
                            "Foundation Key"
                        )
                        == key
                    )
                ]


                top_points_summary = [
                    point
                    for point
                    in st.session_state
                    .foundation_top_points
                    if (
                        point.get(
                            "Foundation Key"
                        )
                        == key
                    )
                ]


                summary_rows.append(
                    {
                        "Project":
                            record.get(
                                "Project",
                                "",
                            ),

                        "Plot / House":
                            record.get(
                                "Plot / House",
                                "",
                            ),

                        "Started":
                            format_date_for_display(
                                record.get(
                                    "Started",
                                    "",
                                )
                            ),

                        "Last Updated":
                            format_date_for_display(
                                record.get(
                                    "Last Updated",
                                    "",
                                )
                            ),

                        "Progress":
                            (
                                f"{calculate_foundation_progress(record)}%"
                            ),

                        "Status":
                            foundation_overall_status(
                                record,
                                bottom_points_summary,
                                top_points_summary,
                            ),
                    }
                )


            st.subheader(
                "Existing Foundation Checklists"
            )

            st.dataframe(
                pd.DataFrame(
                    summary_rows
                ),
                use_container_width=True,
                hide_index=True,
            )


            selected_key = (
                st.selectbox(
                    "Select Foundation Checklist",
                    existing_keys,
                    format_func=lambda key: (
                        f"{key} — "
                        f"{calculate_foundation_progress(st.session_state.foundation_checklists[key])}%"
                    ),
                )
            )


            saved_record = (
                st.session_state
                .foundation_checklists[
                    selected_key
                ]
            )

            current_key = (
                selected_key
            )

            foundation_project = (
                saved_record.get(
                    "Project",
                    "",
                )
            )

            foundation_plot = (
                saved_record.get(
                    "Plot / House",
                    "",
                )
            )

            foundation_date = (
                saved_record.get(
                    "Started",
                    date.today(),
                )
            )

            suffix = (
                widget_suffix(
                    current_key
                )
            )


            st.success(
                (
                    f"Opened: "
                    f"{foundation_project} | "
                    f"{foundation_plot}"
                )
            )


            info1, info2, info3 = (
                st.columns(3)
            )

            with info1:

                st.metric(
                    "Started",
                    format_date_for_display(
                        saved_record.get(
                            "Started",
                            "-",
                        )
                    ),
                )

            with info2:

                st.metric(
                    "Last Updated",
                    format_date_for_display(
                        saved_record.get(
                            "Last Updated",
                            "-",
                        )
                    ),
                )

            with info3:

                st.metric(
                    "Current Progress",
                    (
                        f"{calculate_foundation_progress(saved_record)}%"
                    ),
                )


        else:

            saved_record = {}

            foundation_project = ""

            foundation_plot = ""

            foundation_date = (
                date.today()
            )

            current_key = None

            suffix = (
                "NEW_FOUNDATION"
            )


        st.divider()

        st.subheader(
            "Foundation Identification"
        )


        h1, h2, h3 = (
            st.columns(3)
        )

        with h1:

            foundation_date_input = (
                st.date_input(
                    "Started Date",
                    value=foundation_date,
                    disabled=(
                        foundation_mode
                        == "📂 Open Existing Checklist"
                    ),
                    key=(
                        f"foundation_date_"
                        f"{suffix}"
                    ),
                )
            )

        with h2:

            foundation_project_input = (
                st.text_input(
                    "Project / Site",
                    value=foundation_project,
                    placeholder="Ex.: M1.02",
                    disabled=(
                        foundation_mode
                        == "📂 Open Existing Checklist"
                    ),
                    key=(
                        f"foundation_project_"
                        f"{suffix}"
                    ),
                )
            )

        with h3:

            foundation_plot_input = (
                st.text_input(
                    "Plot / House",
                    value=foundation_plot,
                    placeholder="Ex.: BT05",
                    disabled=(
                        foundation_mode
                        == "📂 Open Existing Checklist"
                    ),
                    key=(
                        f"foundation_plot_"
                        f"{suffix}"
                    ),
                )
            )


        if (
            foundation_mode
            == "➕ New Foundation Checklist"
        ):

            working_key = (
                foundation_key(
                    foundation_project_input,
                    foundation_plot_input,
                )
            )

        else:

            working_key = (
                current_key
            )


        # ====================================================
        # 1 SETTING OUT
        # ====================================================

        st.divider()

        st.subheader(
            "1. Excavation Setting Out"
        )


        status_options = [
            "Not Started",
            "In Progress",
            "Completed",
            "N/A",
        ]


        s1, s2 = (
            st.columns(2)
        )

        with s1:

            setting_position = (
                st.selectbox(
                    "Excavation Position Setting Out",
                    status_options,
                    index=safe_index(
                        status_options,
                        saved_record.get(
                            "Excavation Position Setting Out",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"setting_position_"
                        f"{suffix}"
                    ),
                )
            )

        with s2:

            setting_level = (
                st.selectbox(
                    "Excavation Level Setting Out",
                    status_options,
                    index=safe_index(
                        status_options,
                        saved_record.get(
                            "Excavation Level Setting Out",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"setting_level_"
                        f"{suffix}"
                    ),
                )
            )


        # ====================================================
        # 2 EXCAVATION
        # ====================================================

        st.divider()

        st.subheader(
            "2. Excavation"
        )


        process_options = [
            "Not Started",
            "In Progress",
            "Completed",
        ]


        e1, e2 = (
            st.columns(2)
        )

        with e1:

            excavation_completed = (
                st.selectbox(
                    "Excavation Completed",
                    process_options,
                    index=safe_index(
                        process_options,
                        saved_record.get(
                            "Excavation Completed",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"excavation_completed_"
                        f"{suffix}"
                    ),
                )
            )

        with e2:

            bottom_survey_completed = (
                st.selectbox(
                    "Bottom of Foundation Survey Completed",
                    process_options,
                    index=safe_index(
                        process_options,
                        saved_record.get(
                            "Bottom Survey Completed",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"bottom_survey_completed_"
                        f"{suffix}"
                    ),
                )
            )


        e3, e4 = (
            st.columns(2)
        )

        with e3:

            step_required = (
                st.selectbox(
                    "Foundation Step Required?",
                    [
                        "No",
                        "Yes",
                    ],
                    index=safe_index(
                        [
                            "No",
                            "Yes",
                        ],
                        saved_record.get(
                            "Foundation Step Required",
                            "No",
                        ),
                        "No",
                    ),
                    key=(
                        f"step_required_"
                        f"{suffix}"
                    ),
                )
            )

        with e4:

            if (
                step_required
                == "Yes"
            ):

                step_correct = (
                    st.selectbox(
                        "Foundation Step Checked and Correct?",
                        [
                            "Yes",
                            "No",
                        ],
                        index=safe_index(
                            [
                                "Yes",
                                "No",
                            ],
                            saved_record.get(
                                "Foundation Step Correct",
                                "Yes",
                            ),
                            "Yes",
                        ),
                        key=(
                            f"step_correct_"
                            f"{suffix}"
                        ),
                    )
                )

            else:

                step_correct = (
                    "N/A"
                )

                st.info(
                    "Foundation Step: N/A"
                )


        e5, e6, e7 = (
            st.columns(3)
        )

        with e5:

            no_water = (
                st.selectbox(
                    "Foundation Free of Standing Water?",
                    [
                        "Yes",
                        "No",
                    ],
                    index=safe_index(
                        [
                            "Yes",
                            "No",
                        ],
                        saved_record.get(
                            "No Standing Water",
                            "Yes",
                        ),
                        "Yes",
                    ),
                    key=(
                        f"no_water_"
                        f"{suffix}"
                    ),
                )
            )

        with e6:

            excavation_clean = (
                st.selectbox(
                    "Excavation Bottom Clean?",
                    [
                        "Yes",
                        "No",
                    ],
                    index=safe_index(
                        [
                            "Yes",
                            "No",
                        ],
                        saved_record.get(
                            "Excavation Bottom Clean",
                            "Yes",
                        ),
                        "Yes",
                    ),
                    key=(
                        f"excavation_clean_"
                        f"{suffix}"
                    ),
                )
            )

        with e7:

            no_loose_material = (
                st.selectbox(
                    "No Loose / Fallen Material?",
                    [
                        "Yes",
                        "No",
                    ],
                    index=safe_index(
                        [
                            "Yes",
                            "No",
                        ],
                        saved_record.get(
                            "No Loose Material",
                            "Yes",
                        ),
                        "Yes",
                    ),
                    key=(
                        f"no_loose_material_"
                        f"{suffix}"
                    ),
                )
            )


        # ====================================================
        # BOTTOM SURVEY
        # ====================================================

        st.markdown(
            "### Bottom of Foundation Survey"
        )


        b1, b2 = (
            st.columns(2)
        )

        with b1:

            bottom_design_level = (
                st.number_input(
                    "Design Bottom of Foundation Elevation (m)",
                    value=float(
                        saved_record.get(
                            "Design Bottom Elevation",
                            0.000,
                        )
                    ),
                    step=0.001,
                    format="%.3f",
                    key=(
                        f"bottom_design_"
                        f"{suffix}"
                    ),
                )
            )

        with b2:

            bottom_tolerance = (
                st.number_input(
                    "Survey Tolerance (± mm)",
                    min_value=0.0,
                    value=float(
                        saved_record.get(
                            "Bottom Survey Tolerance",
                            8.0,
                        )
                    ),
                    step=1.0,
                    key=(
                        f"bottom_tolerance_"
                        f"{suffix}"
                    ),
                )
            )


        current_bottom_indices = [
            index
            for index, row
            in enumerate(
                st.session_state
                .foundation_bottom_points
            )
            if (
                row.get(
                    "Foundation Key"
                )
                == working_key
            )
        ]


        current_bottom_points = [
            st.session_state
            .foundation_bottom_points[
                index
            ]
            for index
            in current_bottom_indices
        ]


        if (
            working_key
            and
            working_key
            != " | "
        ):

            with st.form(
                f"bottom_form_{suffix}",
                clear_on_submit=True,
            ):

                bp1, bp2 = (
                    st.columns(2)
                )

                with bp1:

                    bottom_point_name = (
                        st.text_input(
                            "Point / Reference",
                            placeholder="Ex.: P01",
                        )
                    )

                with bp2:

                    bottom_measured_level = (
                        st.number_input(
                            "GPS Surveyed Elevation (m)",
                            value=0.000,
                            step=0.001,
                            format="%.3f",
                        )
                    )

                bottom_point_notes = (
                    st.text_input(
                        "Notes",
                        placeholder="Optional",
                    )
                )

                add_bottom_point = (
                    st.form_submit_button(
                        "Add Bottom Survey Point"
                    )
                )


            if add_bottom_point:

                if (
                    not foundation_project_input.strip()
                    or
                    not foundation_plot_input.strip()
                ):

                    st.error(
                        "Enter Project / Site and Plot / House first."
                    )

                else:

                    bottom_difference = (
                        difference_mm(
                            bottom_measured_level,
                            bottom_design_level,
                        )
                    )

                    bottom_status = (
                        check_tolerance(
                            bottom_difference,
                            bottom_tolerance,
                        )
                    )

                    st.session_state.foundation_bottom_points.append(
                        {
                            "Foundation Key":
                                working_key,

                            "Project":
                                foundation_project_input,

                            "Plot / House":
                                foundation_plot_input,

                            "Date":
                                date.today(),

                            "Point":
                                bottom_point_name,

                            "Design Elevation (m)":
                                bottom_design_level,

                            "Surveyed Elevation (m)":
                                bottom_measured_level,

                            "Difference (mm)":
                                round(
                                    bottom_difference,
                                    1,
                                ),

                            "Tolerance (mm)":
                                bottom_tolerance,

                            "Status":
                                bottom_status,

                            "Notes":
                                bottom_point_notes,
                        }
                    )

                    st.rerun()


        if current_bottom_points:

            df_bottom = (
                pd.DataFrame(
                    current_bottom_points
                )
            )

            display_bottom = (
                df_bottom.copy()
            )

            display_bottom[
                "Status"
            ] = (
                display_bottom[
                    "Status"
                ].apply(
                    lambda value:
                    "✅ PASS"
                    if value == "PASS"
                    else "⚠️ NO PASS"
                )
            )

            st.dataframe(
                display_bottom,
                use_container_width=True,
                hide_index=True,
            )


            with st.expander(
                "Delete Bottom Survey Point"
            ):

                for (
                    number,
                    original_index,
                ) in enumerate(
                    current_bottom_indices,
                    start=1,
                ):

                    row = (
                        st.session_state
                        .foundation_bottom_points[
                            original_index
                        ]
                    )

                    d1, d2, d3, d4 = (
                        st.columns(
                            [
                                1,
                                3,
                                2,
                                0.7,
                            ]
                        )
                    )

                    with d1:

                        st.write(
                            number
                        )

                    with d2:

                        st.write(
                            row.get(
                                "Point",
                                "Point",
                            )
                        )

                    with d3:

                        st.write(
                            status_icon(
                                row.get(
                                    "Status"
                                )
                            )
                        )

                    with d4:

                        if st.button(
                            "🗑️",
                            key=(
                                f"delete_bottom_"
                                f"{original_index}"
                            ),
                        ):

                            del (
                                st.session_state
                                .foundation_bottom_points[
                                    original_index
                                ]
                            )

                            st.rerun()


        # ====================================================
        # 3 REINFORCEMENT
        # ====================================================

        st.divider()

        st.subheader(
            "3. Reinforcement / Mesh"
        )


        reinforcement_options = [
            "Not Started",
            "In Progress",
            "Completed",
            "N/A",
        ]


        r1, r2 = (
            st.columns(2)
        )

        with r1:

            chairs_installed = (
                st.selectbox(
                    "Concrete Spacers / Chairs Installed",
                    reinforcement_options,
                    index=safe_index(
                        reinforcement_options,
                        saved_record.get(
                            "Concrete Chairs Installed",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"chairs_"
                        f"{suffix}"
                    ),
                )
            )

        with r2:

            mesh_installed = (
                st.selectbox(
                    "Mesh / Reinforcement Installed",
                    reinforcement_options,
                    index=safe_index(
                        reinforcement_options,
                        saved_record.get(
                            "Mesh Installed",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"mesh_"
                        f"{suffix}"
                    ),
                )
            )


        r3, r4 = (
            st.columns(2)
        )

        with r3:

            overlap_checked = (
                st.selectbox(
                    "Mesh Overlap Checked and Correct?",
                    [
                        "Yes",
                        "No",
                        "N/A",
                    ],
                    index=safe_index(
                        [
                            "Yes",
                            "No",
                            "N/A",
                        ],
                        saved_record.get(
                            "Overlap Checked",
                            "Yes",
                        ),
                        "Yes",
                    ),
                    key=(
                        f"overlap_"
                        f"{suffix}"
                    ),
                )
            )

        with r4:

            cover_checked = (
                st.selectbox(
                    "Reinforcement Cover Checked?",
                    [
                        "Yes",
                        "No",
                        "N/A",
                    ],
                    index=safe_index(
                        [
                            "Yes",
                            "No",
                            "N/A",
                        ],
                        saved_record.get(
                            "Cover Checked",
                            "Yes",
                        ),
                        "Yes",
                    ),
                    key=(
                        f"cover_"
                        f"{suffix}"
                    ),
                )
            )


        # ====================================================
        # 4 AKM
        # ====================================================

        st.divider()

        st.subheader(
            "4. AKM Inspection"
        )


        yes_no_na = [
            "Yes",
            "No",
            "N/A",
        ]


        i1, i2, i3 = (
            st.columns(3)
        )

        with i1:

            akm_requested = (
                st.selectbox(
                    "AKM Inspection Requested?",
                    yes_no_na,
                    index=safe_index(
                        yes_no_na,
                        saved_record.get(
                            "AKM Inspection Requested",
                            "No",
                        ),
                        "No",
                    ),
                    key=(
                        f"akm_requested_"
                        f"{suffix}"
                    ),
                )
            )

        with i2:

            akm_completed = (
                st.selectbox(
                    "AKM Inspection Completed?",
                    yes_no_na,
                    index=safe_index(
                        yes_no_na,
                        saved_record.get(
                            "AKM Inspection Completed",
                            "No",
                        ),
                        "No",
                    ),
                    key=(
                        f"akm_completed_"
                        f"{suffix}"
                    ),
                )
            )


        akm_result_options = [
            "Pending",
            "Approved",
            "Not Approved",
            "N/A",
        ]


        with i3:

            akm_result = (
                st.selectbox(
                    "AKM Inspection Result",
                    akm_result_options,
                    index=safe_index(
                        akm_result_options,
                        saved_record.get(
                            "AKM Result",
                            "Pending",
                        ),
                        "Pending",
                    ),
                    key=(
                        f"akm_result_"
                        f"{suffix}"
                    ),
                )
            )


        # ====================================================
        # 5 CONCRETE
        # ====================================================

        st.divider()

        st.subheader(
            "5. Concrete"
        )


        concrete_options = [
            "Not Started",
            "In Progress",
            "Completed",
        ]


        c1, c2, c3 = (
            st.columns(3)
        )

        with c1:

            concrete_ordered = (
                st.selectbox(
                    "Concrete Ordered?",
                    concrete_options,
                    index=safe_index(
                        concrete_options,
                        saved_record.get(
                            "Concrete Ordered",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"concrete_ordered_"
                        f"{suffix}"
                    ),
                )
            )

        with c2:

            delivery_confirmed = (
                st.selectbox(
                    "Concrete Delivery Confirmed?",
                    concrete_options,
                    index=safe_index(
                        concrete_options,
                        saved_record.get(
                            "Concrete Delivery Confirmed",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"delivery_confirmed_"
                        f"{suffix}"
                    ),
                )
            )

        with c3:

            concrete_completed = (
                st.selectbox(
                    "Concrete Pour Completed?",
                    concrete_options,
                    index=safe_index(
                        concrete_options,
                        saved_record.get(
                            "Concrete Pour Completed",
                            "Not Started",
                        ),
                        "Not Started",
                    ),
                    key=(
                        f"concrete_completed_"
                        f"{suffix}"
                    ),
                )
            )


        # ====================================================
        # 6 TOP SURVEY
        # ====================================================

        st.divider()

        st.subheader(
            "6. Top of Foundation Survey"
        )


        t1, t2 = (
            st.columns(2)
        )

        with t1:

            foundation_height_mm = (
                st.number_input(
                    "Foundation Design Height (mm)",
                    min_value=0.0,
                    value=float(
                        saved_record.get(
                            "Foundation Height (mm)",
                            300.0,
                        )
                    ),
                    step=10.0,
                    key=(
                        f"foundation_height_"
                        f"{suffix}"
                    ),
                )
            )


        design_top_level = (
            bottom_design_level
            + (
                foundation_height_mm
                / 1000
            )
        )


        with t2:

            st.metric(
                "Calculated Design Top Elevation",
                f"{design_top_level:.3f} m",
            )


        top_survey_completed = (
            st.selectbox(
                "Top of Foundation Survey Completed?",
                process_options,
                index=safe_index(
                    process_options,
                    saved_record.get(
                        "Top Survey Completed",
                        "Not Started",
                    ),
                    "Not Started",
                ),
                key=(
                    f"top_survey_completed_"
                    f"{suffix}"
                ),
            )
        )


        current_top_indices = [
            index
            for index, row
            in enumerate(
                st.session_state
                .foundation_top_points
            )
            if (
                row.get(
                    "Foundation Key"
                )
                == working_key
            )
        ]


        current_top_points = [
            st.session_state
            .foundation_top_points[
                index
            ]
            for index
            in current_top_indices
        ]


        if (
            working_key
            and
            working_key
            != " | "
        ):

            with st.form(
                f"top_form_{suffix}",
                clear_on_submit=True,
            ):

                tp1, tp2 = (
                    st.columns(2)
                )

                with tp1:

                    top_point_name = (
                        st.text_input(
                            "Point / Reference",
                            placeholder="Ex.: P01",
                        )
                    )

                with tp2:

                    top_measured_level = (
                        st.number_input(
                            "GPS Surveyed Top Elevation (m)",
                            value=0.000,
                            step=0.001,
                            format="%.3f",
                        )
                    )

                top_notes = (
                    st.text_input(
                        "Notes",
                        placeholder="Optional",
                    )
                )

                add_top_point = (
                    st.form_submit_button(
                        "Add Top Survey Point"
                    )
                )


            if add_top_point:

                top_difference = (
                    difference_mm(
                        top_measured_level,
                        design_top_level,
                    )
                )

                top_status = (
                    check_tolerance(
                        top_difference,
                        bottom_tolerance,
                    )
                )

                st.session_state.foundation_top_points.append(
                    {
                        "Foundation Key":
                            working_key,

                        "Project":
                            foundation_project_input,

                        "Plot / House":
                            foundation_plot_input,

                        "Date":
                            date.today(),

                        "Point":
                            top_point_name,

                        "Design Elevation (m)":
                            round(
                                design_top_level,
                                3,
                            ),

                        "Surveyed Elevation (m)":
                            top_measured_level,

                        "Difference (mm)":
                            round(
                                top_difference,
                                1,
                            ),

                        "Tolerance (mm)":
                            bottom_tolerance,

                        "Status":
                            top_status,

                        "Notes":
                            top_notes,
                    }
                )

                st.rerun()


        if current_top_points:

            df_top_foundation = (
                pd.DataFrame(
                    current_top_points
                )
            )

            display_top = (
                df_top_foundation.copy()
            )

            display_top[
                "Status"
            ] = (
                display_top[
                    "Status"
                ].apply(
                    lambda value:
                    "✅ PASS"
                    if value == "PASS"
                    else "⚠️ NO PASS"
                )
            )

            st.dataframe(
                display_top,
                use_container_width=True,
                hide_index=True,
            )


            with st.expander(
                "Delete Top Survey Point"
            ):

                for (
                    number,
                    original_index,
                ) in enumerate(
                    current_top_indices,
                    start=1,
                ):

                    row = (
                        st.session_state
                        .foundation_top_points[
                            original_index
                        ]
                    )

                    d1, d2, d3, d4 = (
                        st.columns(
                            [
                                1,
                                3,
                                2,
                                0.7,
                            ]
                        )
                    )

                    with d1:

                        st.write(
                            number
                        )

                    with d2:

                        st.write(
                            row.get(
                                "Point",
                                "Point",
                            )
                        )

                    with d3:

                        st.write(
                            status_icon(
                                row.get(
                                    "Status"
                                )
                            )
                        )

                    with d4:

                        if st.button(
                            "🗑️",
                            key=(
                                f"delete_top_foundation_"
                                f"{original_index}"
                            ),
                        ):

                            del (
                                st.session_state
                                .foundation_top_points[
                                    original_index
                                ]
                            )

                            st.rerun()


        # ====================================================
        # NOTES
        # ====================================================

        st.divider()

        foundation_notes = (
            st.text_area(
                "Foundation Notes / Comments",
                value=saved_record.get(
                    "Notes",
                    "",
                ),
                placeholder=(
                    "Ex.: Waiting for AKM, "
                    "water to be pumped out, "
                    "concrete booked for tomorrow..."
                ),
                key=(
                    f"foundation_notes_"
                    f"{suffix}"
                ),
            )
        )


        # ====================================================
        # SAVE FOUNDATION
        # ====================================================

        button_text = (
            "💾 Create Foundation Checklist"
            if (
                foundation_mode
                == "➕ New Foundation Checklist"
            )
            else
            "💾 Update Foundation Checklist"
        )


        if st.button(
            button_text,
            type="primary",
            key=(
                f"save_foundation_"
                f"{suffix}"
            ),
        ):

            if (
                not foundation_project_input.strip()
                or
                not foundation_plot_input.strip()
            ):

                st.error(
                    "Please enter Project / Site "
                    "and Plot / House."
                )

            else:

                final_key = (
                    foundation_key(
                        foundation_project_input,
                        foundation_plot_input,
                    )
                )


                duplicate = (
                    foundation_mode
                    == "➕ New Foundation Checklist"
                    and
                    final_key
                    in st.session_state
                    .foundation_checklists
                )


                if duplicate:

                    st.error(
                        (
                            "A Foundation Checklist already exists "
                            "for this Project / Plot. "
                            "Use Open Existing Checklist."
                        )
                    )

                else:

                    started_date = (
                        saved_record.get(
                            "Started",
                            foundation_date_input,
                        )
                    )


                    record = {
                        "Started":
                            started_date,

                        "Last Updated":
                            date.today(),

                        "Project":
                            foundation_project_input,

                        "Plot / House":
                            foundation_plot_input,

                        "Excavation Position Setting Out":
                            setting_position,

                        "Excavation Level Setting Out":
                            setting_level,

                        "Excavation Completed":
                            excavation_completed,

                        "Bottom Survey Completed":
                            bottom_survey_completed,

                        "Foundation Step Required":
                            step_required,

                        "Foundation Step Correct":
                            step_correct,

                        "No Standing Water":
                            no_water,

                        "Excavation Bottom Clean":
                            excavation_clean,

                        "No Loose Material":
                            no_loose_material,

                        "Design Bottom Elevation":
                            bottom_design_level,

                        "Bottom Survey Tolerance":
                            bottom_tolerance,

                        "Concrete Chairs Installed":
                            chairs_installed,

                        "Mesh Installed":
                            mesh_installed,

                        "Overlap Checked":
                            overlap_checked,

                        "Cover Checked":
                            cover_checked,

                        "AKM Inspection Requested":
                            akm_requested,

                        "AKM Inspection Completed":
                            akm_completed,

                        "AKM Result":
                            akm_result,

                        "Concrete Ordered":
                            concrete_ordered,

                        "Concrete Delivery Confirmed":
                            delivery_confirmed,

                        "Concrete Pour Completed":
                            concrete_completed,

                        "Foundation Height (mm)":
                            foundation_height_mm,

                        "Design Top Elevation":
                            design_top_level,

                        "Top Survey Completed":
                            top_survey_completed,

                        "Notes":
                            foundation_notes,
                    }


                    st.session_state.foundation_checklists[
                        final_key
                    ] = record


                    if (
                        foundation_mode
                        == "➕ New Foundation Checklist"
                    ):

                        for point in (
                            st.session_state
                            .foundation_bottom_points
                        ):

                            if (
                                point.get(
                                    "Foundation Key"
                                )
                                == working_key
                            ):

                                point[
                                    "Foundation Key"
                                ] = final_key


                        for point in (
                            st.session_state
                            .foundation_top_points
                        ):

                            if (
                                point.get(
                                    "Foundation Key"
                                )
                                == working_key
                            ):

                                point[
                                    "Foundation Key"
                                ] = final_key


                    st.success(
                        (
                            "Foundation Checklist created."
                            if (
                                foundation_mode
                                == "➕ New Foundation Checklist"
                            )
                            else
                            "Foundation Checklist updated."
                        )
                    )

                    st.rerun()


        # ====================================================
        # FOUNDATION STATUS
        # ====================================================

        if (
            foundation_mode
            == "📂 Open Existing Checklist"
        ):

            current_saved_record = (
                st.session_state
                .foundation_checklists[
                    current_key
                ]
            )


            current_bottom_status = [
                point
                for point
                in st.session_state
                .foundation_bottom_points
                if (
                    point.get(
                        "Foundation Key"
                    )
                    == current_key
                )
            ]


            current_top_status = [
                point
                for point
                in st.session_state
                .foundation_top_points
                if (
                    point.get(
                        "Foundation Key"
                    )
                    == current_key
                )
            ]


            progress = (
                calculate_foundation_progress(
                    current_saved_record
                )
            )

            review_items = (
                foundation_review_items(
                    current_saved_record,
                    current_bottom_status,
                    current_top_status,
                )
            )

            overall_status = (
                foundation_overall_status(
                    current_saved_record,
                    current_bottom_status,
                    current_top_status,
                )
            )


            st.divider()

            st.header(
                "Foundation Status"
            )

            st.progress(
                progress / 100
            )

            st.write(
                f"**Progress: {progress}%**"
            )


            o1, o2, o3, o4 = (
                st.columns(4)
            )

            with o1:

                st.metric(
                    "Setting Out",
                    status_icon(
                        current_saved_record.get(
                            "Excavation Position Setting Out",
                            "Not Started",
                        )
                    ),
                )

            with o2:

                st.metric(
                    "AKM",
                    status_icon(
                        current_saved_record.get(
                            "AKM Result",
                            "Pending",
                        )
                    ),
                )

            with o3:

                st.metric(
                    "Concrete",
                    status_icon(
                        current_saved_record.get(
                            "Concrete Pour Completed",
                            "Not Started",
                        )
                    ),
                )

            with o4:

                st.metric(
                    "Top Survey",
                    status_icon(
                        current_saved_record.get(
                            "Top Survey Completed",
                            "Not Started",
                        )
                    ),
                )


            if review_items:

                st.warning(
                    "⚠️ REVIEW REQUIRED"
                )

                review_df = (
                    pd.DataFrame(
                        {
                            "Item Requiring Review":
                                review_items
                        }
                    )
                )

                st.dataframe(
                    review_df,
                    use_container_width=True,
                    hide_index=True,
                )


            if (
                overall_status
                == "COMPLETE"
            ):

                st.success(
                    "✅ FOUNDATION STATUS: COMPLETE"
                )

            elif (
                overall_status
                == "REVIEW REQUIRED"
            ):

                st.warning(
                    "⚠️ FOUNDATION STATUS: "
                    "REVIEW REQUIRED"
                )

            else:

                st.info(
                    "🟡 FOUNDATION STATUS: "
                    "IN PROGRESS"
                )


# ============================================================
# SNAG LIST
# ============================================================

elif pagina == "Snag List":

    st.header(
        "Snag List"
    )

    st.info(
        "Módulo Snag List em construção."
    )


# ============================================================
# AS-BUILT
# ============================================================

elif pagina == "As-Built":

    st.header(
        "As-Built"
    )

    st.info(
        "Módulo As-Built em construção."
    )


# ============================================================
# LEVELS
# ============================================================

elif pagina == "Levels":

    st.header(
        "Levels"
    )

    st.info(
        "Módulo Levels em construção."
    )


# ============================================================
# CONCRETE
# ============================================================

elif pagina == "Concrete":

    st.header(
        "Concrete"
    )

    st.info(
        "Módulo Concrete em construção."
    )


# ============================================================
# DRAINAGE
# ============================================================

elif pagina == "Drainage":

    st.header(
        "Drainage"
    )

    st.info(
        "Módulo Drainage em construção."
    )


# ============================================================
# PUBLIC LIGHTING
# ============================================================

elif pagina == "Public Lighting":

    st.header(
        "Public Lighting"
    )

    st.info(
        "Módulo Public Lighting em construção."
    )


# ============================================================
# DAILY REPORT
# ============================================================

elif pagina == "Daily Report":

    st.header(
        "Daily Report"
    )

    st.caption(
        "Daily site records, manpower, plant, "
        "activities, deliveries, inspections and issues."
    )


    # ========================================================
    # NEW / OPEN DAILY
    # ========================================================

    daily_mode = (
        st.radio(
            "Daily Report Record",
            [
                "➕ New Daily Report",
                "📂 Open Existing Daily Report",
            ],
            horizontal=True,
        )
    )


    existing_daily_keys = list(
        st.session_state
        .daily_reports
        .keys()
    )


    # ========================================================
    # OPEN DAILY
    # ========================================================

    if (
        daily_mode
        == "📂 Open Existing Daily Report"
    ):

        if not existing_daily_keys:

            st.info(
                "No Daily Reports have been created yet."
            )

            st.stop()


        daily_summary_rows = []

        for key in existing_daily_keys:

            record = (
                st.session_state
                .daily_reports[
                    key
                ]
            )


            daily_summary_rows.append(
                {
                    "Project":
                        record.get(
                            "Project",
                            "",
                        ),

                    "Date":
                        format_date_for_display(
                            record.get(
                                "Date",
                                "",
                            )
                        ),

                    "Status":
                        record.get(
                            "Status",
                            "Draft",
                        ),

                    "Manpower":
                        calculate_total_manpower(
                            key
                        ),

                    "Plant":
                        calculate_total_equipment(
                            key
                        ),

                    "Activities":
                        len(
                            daily_items_for_key(
                                st.session_state
                                .daily_activities,
                                key,
                            )
                        ),

                    "Last Updated":
                        format_date_for_display(
                            record.get(
                                "Last Updated",
                                "",
                            )
                        ),
                }
            )


        st.subheader(
            "Existing Daily Reports"
        )

        st.dataframe(
            pd.DataFrame(
                daily_summary_rows
            ),
            use_container_width=True,
            hide_index=True,
        )


        selected_daily_key = (
            st.selectbox(
                "Select Daily Report",
                existing_daily_keys,
                format_func=lambda key: (
                    f"{st.session_state.daily_reports[key].get('Project', '')}"
                    f" | "
                    f"{format_date_for_display(st.session_state.daily_reports[key].get('Date', ''))}"
                    f" | "
                    f"{st.session_state.daily_reports[key].get('Status', 'Draft')}"
                ),
            )
        )


        daily_saved_record = (
            st.session_state
            .daily_reports[
                selected_daily_key
            ]
        )


        daily_current_key = (
            selected_daily_key
        )


        daily_project = (
            daily_saved_record.get(
                "Project",
                "",
            )
        )


        daily_date = (
            daily_saved_record.get(
                "Date",
                date.today(),
            )
        )


        daily_suffix = (
            widget_suffix(
                selected_daily_key
            )
        )


        st.success(
            (
                f"Opened: "
                f"{daily_project} | "
                f"{format_date_for_display(daily_date)}"
            )
        )


        oi1, oi2, oi3 = (
            st.columns(3)
        )

        with oi1:

            st.metric(
                "Report Status",
                status_icon(
                    daily_saved_record.get(
                        "Status",
                        "Draft",
                    )
                ),
            )

        with oi2:

            st.metric(
                "Created",
                format_date_for_display(
                    daily_saved_record.get(
                        "Created",
                        daily_date,
                    )
                ),
            )

        with oi3:

            st.metric(
                "Last Updated",
                format_date_for_display(
                    daily_saved_record.get(
                        "Last Updated",
                        "-",
                    )
                ),
            )


    else:

        daily_saved_record = {}

        daily_project = ""

        daily_date = (
            date.today()
        )

        daily_current_key = None

        daily_suffix = (
            "NEW_DAILY_REPORT"
        )


    st.divider()


    # ========================================================
    # 1 REPORT INFORMATION
    # ========================================================

    st.subheader(
        "1. Report Information"
    )


    ri1, ri2 = (
        st.columns(2)
    )

    with ri1:

        daily_project_input = (
            st.text_input(
                "Project / Site",
                value=daily_project,
                placeholder="Ex.: M1.02",
                disabled=(
                    daily_mode
                    == "📂 Open Existing Daily Report"
                ),
                key=(
                    f"daily_project_"
                    f"{daily_suffix}"
                ),
            )
        )

    with ri2:

        daily_date_input = (
            st.date_input(
                "Report Date",
                value=daily_date,
                disabled=(
                    daily_mode
                    == "📂 Open Existing Daily Report"
                ),
                key=(
                    f"daily_date_"
                    f"{daily_suffix}"
                ),
            )
        )


    ri3, ri4 = (
        st.columns(2)
    )

    with ri3:

        working_start = (
            st.time_input(
                "Working Start Time",
                value=daily_saved_record.get(
                    "Working Start",
                    time(
                        7,
                        30,
                    ),
                ),
                key=(
                    f"daily_start_"
                    f"{daily_suffix}"
                ),
            )
        )

    with ri4:

        working_finish = (
            st.time_input(
                "Working Finish Time",
                value=daily_saved_record.get(
                    "Working Finish",
                    time(
                        17,
                        0,
                    ),
                ),
                key=(
                    f"daily_finish_"
                    f"{daily_suffix}"
                ),
            )
        )


    if (
        daily_mode
        == "➕ New Daily Report"
    ):

        daily_working_key = (
            daily_report_key(
                daily_project_input,
                daily_date_input,
            )
        )

    else:

        daily_working_key = (
            daily_current_key
        )


    # ========================================================
    # 2 MANPOWER
    # ========================================================

    st.divider()

    st.subheader(
        "2. Manpower"
    )

    st.caption(
        "Add each company, role and number of personnel."
    )


    with st.form(
        f"daily_manpower_form_{daily_suffix}",
        clear_on_submit=True,
    ):

        mp1, mp2, mp3 = (
            st.columns(
                [
                    2,
                    2,
                    1,
                ]
            )
        )

        with mp1:

            manpower_company = (
                st.text_input(
                    "Company / Subcontractor",
                    placeholder="Ex.: KEGGALL",
                )
            )

        with mp2:

            manpower_role = (
                st.text_input(
                    "Role / Trade",
                    placeholder="Ex.: Groundworker",
                )
            )

        with mp3:

            manpower_quantity = (
                st.number_input(
                    "Quantity",
                    min_value=1,
                    value=1,
                    step=1,
                )
            )


        add_manpower = (
            st.form_submit_button(
                "Add Manpower"
            )
        )


    if add_manpower:

        if (
            not daily_project_input.strip()
        ):

            st.error(
                "Enter Project / Site first."
            )

        elif (
            not manpower_company.strip()
            or
            not manpower_role.strip()
        ):

            st.error(
                "Enter Company and Role."
            )

        else:

            st.session_state.daily_manpower.append(
                {
                    "Daily Report Key":
                        daily_working_key,

                    "Company":
                        manpower_company.strip(),

                    "Role":
                        manpower_role.strip(),

                    "Quantity":
                        int(
                            manpower_quantity
                        ),
                }
            )

            st.rerun()


    current_manpower_indices = [
        index
        for index, item
        in enumerate(
            st.session_state
            .daily_manpower
        )
        if (
            item.get(
                "Daily Report Key"
            )
            == daily_working_key
        )
    ]


    current_manpower = [
        st.session_state
        .daily_manpower[
            index
        ]
        for index
        in current_manpower_indices
    ]


    if current_manpower:

        manpower_display = (
            pd.DataFrame(
                current_manpower
            )
        )

        manpower_display = (
            manpower_display[
                [
                    "Company",
                    "Role",
                    "Quantity",
                ]
            ]
        )


        st.dataframe(
            manpower_display,
            use_container_width=True,
            hide_index=True,
        )


        st.metric(
            "Total Manpower",
            calculate_total_manpower(
                daily_working_key
            ),
        )


        with st.expander(
            "Edit / Delete Manpower"
        ):

            for (
                number,
                original_index,
            ) in enumerate(
                current_manpower_indices,
                start=1,
            ):

                item = (
                    st.session_state
                    .daily_manpower[
                        original_index
                    ]
                )


                dm1, dm2, dm3, dm4 = (
                    st.columns(
                        [
                            1,
                            3,
                            2,
                            0.7,
                        ]
                    )
                )

                with dm1:

                    st.write(
                        number
                    )

                with dm2:

                    st.write(
                        (
                            f"{item.get('Company', '')} - "
                            f"{item.get('Role', '')}"
                        )
                    )

                with dm3:

                    st.write(
                        (
                            f"Qty: "
                            f"{item.get('Quantity', 0)}"
                        )
                    )

                with dm4:

                    if st.button(
                        "🗑️",
                        key=(
                            f"delete_manpower_"
                            f"{original_index}"
                        ),
                    ):

                        del (
                            st.session_state
                            .daily_manpower[
                                original_index
                            ]
                        )

                        st.rerun()


    # ========================================================
    # 3 PLANT
    # ========================================================

    st.divider()

    st.subheader(
        "3. Plant & Equipment"
    )


    with st.form(
        f"daily_equipment_form_{daily_suffix}",
        clear_on_submit=True,
    ):

        eq1, eq2, eq3 = (
            st.columns(
                [
                    2,
                    2,
                    1,
                ]
            )
        )

        with eq1:

            equipment_name = (
                st.text_input(
                    "Equipment",
                    placeholder="Ex.: Excavator 14T",
                )
            )

        with eq2:

            equipment_company = (
                st.text_input(
                    "Company / Owner",
                    placeholder="Optional",
                )
            )

        with eq3:

            equipment_quantity = (
                st.number_input(
                    "Quantity",
                    min_value=1,
                    value=1,
                    step=1,
                )
            )


        add_equipment = (
            st.form_submit_button(
                "Add Equipment"
            )
        )


    if add_equipment:

        if (
            not equipment_name.strip()
        ):

            st.error(
                "Enter the equipment description."
            )

        else:

            st.session_state.daily_equipment.append(
                {
                    "Daily Report Key":
                        daily_working_key,

                    "Equipment":
                        equipment_name.strip(),

                    "Company / Owner":
                        equipment_company.strip(),

                    "Quantity":
                        int(
                            equipment_quantity
                        ),
                }
            )

            st.rerun()


    current_equipment_indices = [
        index
        for index, item
        in enumerate(
            st.session_state
            .daily_equipment
        )
        if (
            item.get(
                "Daily Report Key"
            )
            == daily_working_key
        )
    ]


    current_equipment = [
        st.session_state
        .daily_equipment[
            index
        ]
        for index
        in current_equipment_indices
    ]


    if current_equipment:

        equipment_display = (
            pd.DataFrame(
                current_equipment
            )
        )

        equipment_display = (
            equipment_display[
                [
                    "Equipment",
                    "Company / Owner",
                    "Quantity",
                ]
            ]
        )


        st.dataframe(
            equipment_display,
            use_container_width=True,
            hide_index=True,
        )


        st.metric(
            "Total Plant / Equipment",
            calculate_total_equipment(
                daily_working_key
            ),
        )


        with st.expander(
            "Edit / Delete Equipment"
        ):

            for (
                number,
                original_index,
            ) in enumerate(
                current_equipment_indices,
                start=1,
            ):

                item = (
                    st.session_state
                    .daily_equipment[
                        original_index
                    ]
                )


                de1, de2, de3, de4 = (
                    st.columns(
                        [
                            1,
                            3,
                            2,
                            0.7,
                        ]
                    )
                )

                with de1:

                    st.write(
                        number
                    )

                with de2:

                    st.write(
                        item.get(
                            "Equipment",
                            "",
                        )
                    )

                with de3:

                    st.write(
                        (
                            f"Qty: "
                            f"{item.get('Quantity', 0)}"
                        )
                    )

                with de4:

                    if st.button(
                        "🗑️",
                        key=(
                            f"delete_equipment_"
                            f"{original_index}"
                        ),
                    ):

                        del (
                            st.session_state
                            .daily_equipment[
                                original_index
                            ]
                        )

                        st.rerun()


    # ========================================================
    # 4 WEATHER
    # ========================================================

    st.divider()

    st.subheader(
        "4. Weather & Workability"
    )


    weather_options = [
        "Good / Dry",
        "Cloudy",
        "Rainy",
        "Heavy Rain",
        "Windy",
        "Other",
    ]


    workability_options = [
        "Workable",
        "Partially Workable",
        "Not Workable",
    ]


    w1, w2 = (
        st.columns(2)
    )

    with w1:

        weather_morning = (
            st.selectbox(
                "Morning Weather",
                weather_options,
                index=safe_index(
                    weather_options,
                    daily_saved_record.get(
                        "Weather Morning",
                        "Good / Dry",
                    ),
                    "Good / Dry",
                ),
                key=(
                    f"weather_morning_"
                    f"{daily_suffix}"
                ),
            )
        )

    with w2:

        workability_morning = (
            st.selectbox(
                "Morning Workability",
                workability_options,
                index=safe_index(
                    workability_options,
                    daily_saved_record.get(
                        "Workability Morning",
                        "Workable",
                    ),
                    "Workable",
                ),
                key=(
                    f"workability_morning_"
                    f"{daily_suffix}"
                ),
            )
        )


    w3, w4 = (
        st.columns(2)
    )

    with w3:

        weather_afternoon = (
            st.selectbox(
                "Afternoon Weather",
                weather_options,
                index=safe_index(
                    weather_options,
                    daily_saved_record.get(
                        "Weather Afternoon",
                        "Good / Dry",
                    ),
                    "Good / Dry",
                ),
                key=(
                    f"weather_afternoon_"
                    f"{daily_suffix}"
                ),
            )
        )

    with w4:

        workability_afternoon = (
            st.selectbox(
                "Afternoon Workability",
                workability_options,
                index=safe_index(
                    workability_options,
                    daily_saved_record.get(
                        "Workability Afternoon",
                        "Workable",
                    ),
                    "Workable",
                ),
                key=(
                    f"workability_afternoon_"
                    f"{daily_suffix}"
                ),
            )
        )


    weather_notes = (
        st.text_input(
            "Weather Notes",
            value=daily_saved_record.get(
                "Weather Notes",
                "",
            ),
            placeholder="Optional",
            key=(
                f"weather_notes_"
                f"{daily_suffix}"
            ),
        )
    )


    # ========================================================
    # 5 WORK PERFORMED
    # ========================================================

    st.divider()

    st.subheader(
        "5. Work Performed"
    )

    st.caption(
        "Record each activity separately by location."
    )


    with st.form(
        f"daily_activity_form_{daily_suffix}",
        clear_on_submit=True,
    ):

        ac1, ac2 = (
            st.columns(2)
        )

        with ac1:

            activity_location = (
                st.text_input(
                    "Location / Plot",
                    placeholder="Ex.: BT05",
                )
            )

        with ac2:

            activity_company = (
                st.text_input(
                    "Company / Subcontractor",
                    placeholder="Optional",
                )
            )


        activity_description = (
            st.text_area(
                "Activity / Work Performed",
                placeholder=(
                    "Ex.: Adjustment of chamber levels"
                ),
            )
        )


        activity_notes = (
            st.text_input(
                "Activity Notes",
                placeholder="Optional",
            )
        )


        add_activity = (
            st.form_submit_button(
                "Add Activity"
            )
        )


    if add_activity:

        if (
            not activity_description.strip()
        ):

            st.error(
                "Enter the activity performed."
            )

        else:

            st.session_state.daily_activities.append(
                {
                    "Daily Report Key":
                        daily_working_key,

                    "Location / Plot":
                        activity_location.strip(),

                    "Company":
                        activity_company.strip(),

                    "Activity":
                        activity_description.strip(),

                    "Notes":
                        activity_notes.strip(),
                }
            )

            st.rerun()


    current_activity_indices = [
        index
        for index, item
        in enumerate(
            st.session_state
            .daily_activities
        )
        if (
            item.get(
                "Daily Report Key"
            )
            == daily_working_key
        )
    ]


    current_activities = [
        st.session_state
        .daily_activities[
            index
        ]
        for index
        in current_activity_indices
    ]


    if current_activities:

        activity_display = (
            pd.DataFrame(
                current_activities
            )
        )

        activity_display = (
            activity_display[
                [
                    "Location / Plot",
                    "Company",
                    "Activity",
                    "Notes",
                ]
            ]
        )


        st.dataframe(
            activity_display,
            use_container_width=True,
            hide_index=True,
        )


        with st.expander(
            "Edit / Delete Activities"
        ):

            for (
                number,
                original_index,
            ) in enumerate(
                current_activity_indices,
                start=1,
            ):

                item = (
                    st.session_state
                    .daily_activities[
                        original_index
                    ]
                )


                da1, da2, da3, da4 = (
                    st.columns(
                        [
                            1,
                            2,
                            5,
                            0.7,
                        ]
                    )
                )

                with da1:

                    st.write(
                        number
                    )

                with da2:

                    st.write(
                        item.get(
                            "Location / Plot",
                            "",
                        )
                    )

                with da3:

                    st.write(
                        item.get(
                            "Activity",
                            "",
                        )
                    )

                with da4:

                    if st.button(
                        "🗑️",
                        key=(
                            f"delete_activity_"
                            f"{original_index}"
                        ),
                    ):

                        del (
                            st.session_state
                            .daily_activities[
                                original_index
                            ]
                        )

                        st.rerun()


    # ========================================================
    # 6 DELIVERIES
    # ========================================================

    st.divider()

    st.subheader(
        "6. Deliveries / Materials"
    )


    with st.form(
        f"daily_delivery_form_{daily_suffix}",
        clear_on_submit=True,
    ):

        dl1, dl2 = (
            st.columns(2)
        )

        with dl1:

            delivery_material = (
                st.text_input(
                    "Material / Delivery",
                    placeholder="Ex.: Stone",
                )
            )

        with dl2:

            delivery_quantity = (
                st.text_input(
                    "Quantity",
                    placeholder=(
                        "Ex.: 20 t / 12 m³ / 3 loads"
                    ),
                )
            )


        dl3, dl4 = (
            st.columns(2)
        )

        with dl3:

            delivery_supplier = (
                st.text_input(
                    "Supplier",
                    placeholder="Optional",
                )
            )

        with dl4:

            delivery_location = (
                st.text_input(
                    "Delivery Location",
                    placeholder="Ex.: BT06",
                )
            )


        add_delivery = (
            st.form_submit_button(
                "Add Delivery"
            )
        )


    if add_delivery:

        if (
            not delivery_material.strip()
        ):

            st.error(
                "Enter the material / delivery."
            )

        else:

            st.session_state.daily_deliveries.append(
                {
                    "Daily Report Key":
                        daily_working_key,

                    "Material / Delivery":
                        delivery_material.strip(),

                    "Quantity":
                        delivery_quantity.strip(),

                    "Supplier":
                        delivery_supplier.strip(),

                    "Location":
                        delivery_location.strip(),
                }
            )

            st.rerun()


    current_delivery_indices = [
        index
        for index, item
        in enumerate(
            st.session_state
            .daily_deliveries
        )
        if (
            item.get(
                "Daily Report Key"
            )
            == daily_working_key
        )
    ]


    current_deliveries = [
        st.session_state
        .daily_deliveries[
            index
        ]
        for index
        in current_delivery_indices
    ]


    if current_deliveries:

        delivery_display = (
            pd.DataFrame(
                current_deliveries
            )
        )

        delivery_display = (
            delivery_display[
                [
                    "Material / Delivery",
                    "Quantity",
                    "Supplier",
                    "Location",
                ]
            ]
        )


        st.dataframe(
            delivery_display,
            use_container_width=True,
            hide_index=True,
        )


        with st.expander(
            "Edit / Delete Deliveries"
        ):

            for (
                number,
                original_index,
            ) in enumerate(
                current_delivery_indices,
                start=1,
            ):

                item = (
                    st.session_state
                    .daily_deliveries[
                        original_index
                    ]
                )


                dd1, dd2, dd3, dd4 = (
                    st.columns(
                        [
                            1,
                            4,
                            2,
                            0.7,
                        ]
                    )
                )

                with dd1:

                    st.write(
                        number
                    )

                with dd2:

                    st.write(
                        item.get(
                            "Material / Delivery",
                            "",
                        )
                    )

                with dd3:

                    st.write(
                        item.get(
                            "Quantity",
                            "",
                        )
                    )

                with dd4:

                    if st.button(
                        "🗑️",
                        key=(
                            f"delete_delivery_"
                            f"{original_index}"
                        ),
                    ):

                        del (
                            st.session_state
                            .daily_deliveries[
                                original_index
                            ]
                        )

                        st.rerun()


    # ========================================================
    # 7 INSPECTIONS
    # ========================================================

    st.divider()

    st.subheader(
        "7. Inspections / Surveys"
    )


    with st.form(
        f"daily_inspection_form_{daily_suffix}",
        clear_on_submit=True,
    ):

        ins1, ins2, ins3 = (
            st.columns(3)
        )

        with ins1:

            inspection_type = (
                st.text_input(
                    "Inspection / Survey Type",
                    placeholder="Ex.: AKM Inspection",
                )
            )

        with ins2:

            inspection_location = (
                st.text_input(
                    "Location",
                    placeholder="Ex.: BT05",
                )
            )

        with ins3:

            inspection_status = (
                st.selectbox(
                    "Status",
                    [
                        "Pending",
                        "Completed",
                        "Approved",
                        "Not Approved",
                    ],
                )
            )


        inspection_notes = (
            st.text_input(
                "Inspection Notes",
                placeholder="Optional",
            )
        )


        add_inspection = (
            st.form_submit_button(
                "Add Inspection / Survey"
            )
        )


    if add_inspection:

        if (
            not inspection_type.strip()
        ):

            st.error(
                "Enter the inspection / survey type."
            )

        else:

            st.session_state.daily_inspections.append(
                {
                    "Daily Report Key":
                        daily_working_key,

                    "Type":
                        inspection_type.strip(),

                    "Location":
                        inspection_location.strip(),

                    "Status":
                        inspection_status,

                    "Notes":
                        inspection_notes.strip(),
                }
            )

            st.rerun()


    current_inspection_indices = [
        index
        for index, item
        in enumerate(
            st.session_state
            .daily_inspections
        )
        if (
            item.get(
                "Daily Report Key"
            )
            == daily_working_key
        )
    ]


    current_inspections = [
        st.session_state
        .daily_inspections[
            index
        ]
        for index
        in current_inspection_indices
    ]


    if current_inspections:

        inspection_display = (
            pd.DataFrame(
                current_inspections
            )
        )

        inspection_display = (
            inspection_display[
                [
                    "Type",
                    "Location",
                    "Status",
                    "Notes",
                ]
            ]
        )


        st.dataframe(
            inspection_display,
            use_container_width=True,
            hide_index=True,
        )


        with st.expander(
            "Edit / Delete Inspections"
        ):

            for (
                number,
                original_index,
            ) in enumerate(
                current_inspection_indices,
                start=1,
            ):

                item = (
                    st.session_state
                    .daily_inspections[
                        original_index
                    ]
                )


                di1, di2, di3, di4 = (
                    st.columns(
                        [
                            1,
                            4,
                            2,
                            0.7,
                        ]
                    )
                )

                with di1:

                    st.write(
                        number
                    )

                with di2:

                    st.write(
                        item.get(
                            "Type",
                            "",
                        )
                    )

                with di3:

                    st.write(
                        item.get(
                            "Status",
                            "",
                        )
                    )

                with di4:

                    if st.button(
                        "🗑️",
                        key=(
                            f"delete_inspection_"
                            f"{original_index}"
                        ),
                    ):

                        del (
                            st.session_state
                            .daily_inspections[
                                original_index
                            ]
                        )

                        st.rerun()


    # ========================================================
    # 8 ISSUES
    # ========================================================

    st.divider()

    st.subheader(
        "8. Issues / Delays / Constraints"
    )


    with st.form(
        f"daily_issue_form_{daily_suffix}",
        clear_on_submit=True,
    ):

        issue_description = (
            st.text_area(
                "Issue / Delay",
                placeholder=(
                    "Ex.: Standing water in excavation"
                ),
            )
        )


        is1, is2 = (
            st.columns(2)
        )

        with is1:

            issue_location = (
                st.text_input(
                    "Location",
                    placeholder="Ex.: BT05",
                )
            )

        with is2:

            issue_impact = (
                st.text_input(
                    "Impact",
                    placeholder=(
                        "Ex.: Foundation preparation delayed"
                    ),
                )
            )


        issue_action = (
            st.text_input(
                "Action / Follow-up",
                placeholder=(
                    "Ex.: Pump out required before reinforcement"
                ),
            )
        )


        add_issue = (
            st.form_submit_button(
                "Add Issue"
            )
        )


    if add_issue:

        if (
            not issue_description.strip()
        ):

            st.error(
                "Enter the issue / delay."
            )

        else:

            st.session_state.daily_issues.append(
                {
                    "Daily Report Key":
                        daily_working_key,

                    "Issue":
                        issue_description.strip(),

                    "Location":
                        issue_location.strip(),

                    "Impact":
                        issue_impact.strip(),

                    "Action / Follow-up":
                        issue_action.strip(),
                }
            )

            st.rerun()


    current_issue_indices = [
        index
        for index, item
        in enumerate(
            st.session_state
            .daily_issues
        )
        if (
            item.get(
                "Daily Report Key"
            )
            == daily_working_key
        )
    ]


    current_issues = [
        st.session_state
        .daily_issues[
            index
        ]
        for index
        in current_issue_indices
    ]


    if current_issues:

        issue_display = (
            pd.DataFrame(
                current_issues
            )
        )

        issue_display = (
            issue_display[
                [
                    "Issue",
                    "Location",
                    "Impact",
                    "Action / Follow-up",
                ]
            ]
        )


        st.dataframe(
            issue_display,
            use_container_width=True,
            hide_index=True,
        )


        with st.expander(
            "Edit / Delete Issues"
        ):

            for (
                number,
                original_index,
            ) in enumerate(
                current_issue_indices,
                start=1,
            ):

                item = (
                    st.session_state
                    .daily_issues[
                        original_index
                    ]
                )


                ds1, ds2, ds3, ds4 = (
                    st.columns(
                        [
                            1,
                            5,
                            2,
                            0.7,
                        ]
                    )
                )

                with ds1:

                    st.write(
                        number
                    )

                with ds2:

                    st.write(
                        item.get(
                            "Issue",
                            "",
                        )
                    )

                with ds3:

                    st.write(
                        item.get(
                            "Location",
                            "",
                        )
                    )

                with ds4:

                    if st.button(
                        "🗑️",
                        key=(
                            f"delete_issue_"
                            f"{original_index}"
                        ),
                    ):

                        del (
                            st.session_state
                            .daily_issues[
                                original_index
                            ]
                        )

                        st.rerun()


    # ========================================================
    # 9 PHOTOS
    # ========================================================

    st.divider()

    st.subheader(
        "9. Photos"
    )


    uploaded_photos = (
        st.file_uploader(
            "Upload Site Photos",
            type=[
                "jpg",
                "jpeg",
                "png",
            ],
            accept_multiple_files=True,
            key=(
                f"daily_photos_upload_"
                f"{daily_suffix}"
            ),
        )
    )


    if uploaded_photos:

        existing_photo_names = {
            photo.get(
                "File Name"
            )
            for photo
            in st.session_state
            .daily_photos
            if (
                photo.get(
                    "Daily Report Key"
                )
                == daily_working_key
            )
        }


        for uploaded_file in (
            uploaded_photos
        ):

            if (
                uploaded_file.name
                not in existing_photo_names
            ):

                st.session_state.daily_photos.append(
                    {
                        "Daily Report Key":
                            daily_working_key,

                        "File Name":
                            uploaded_file.name,

                        "Bytes":
                            uploaded_file.getvalue(),

                        "Location":
                            "",

                        "Description":
                            "",
                    }
                )

                existing_photo_names.add(
                    uploaded_file.name
                )


    current_photo_indices = [
        index
        for index, item
        in enumerate(
            st.session_state
            .daily_photos
        )
        if (
            item.get(
                "Daily Report Key"
            )
            == daily_working_key
        )
    ]


    current_photos = [
        st.session_state
        .daily_photos[
            index
        ]
        for index
        in current_photo_indices
    ]


    if current_photos:

        st.write(
            (
                f"**Photos recorded: "
                f"{len(current_photos)}**"
            )
        )


        for (
            number,
            original_index,
        ) in enumerate(
            current_photo_indices,
            start=1,
        ):

            photo = (
                st.session_state
                .daily_photos[
                    original_index
                ]
            )


            with st.expander(
                (
                    f"Photo {number} - "
                    f"{photo.get('File Name', '')}"
                )
            ):

                st.image(
                    photo.get(
                        "Bytes"
                    ),
                    width=500,
                )


                pc1, pc2 = (
                    st.columns(2)
                )

                with pc1:

                    photo_location = (
                        st.text_input(
                            "Photo Location",
                            value=photo.get(
                                "Location",
                                "",
                            ),
                            key=(
                                f"photo_location_"
                                f"{original_index}"
                            ),
                        )
                    )

                with pc2:

                    photo_description = (
                        st.text_input(
                            "Photo Description",
                            value=photo.get(
                                "Description",
                                "",
                            ),
                            key=(
                                f"photo_description_"
                                f"{original_index}"
                            ),
                        )
                    )


                st.session_state.daily_photos[
                    original_index
                ][
                    "Location"
                ] = photo_location


                st.session_state.daily_photos[
                    original_index
                ][
                    "Description"
                ] = photo_description


                if st.button(
                    "🗑️ Delete Photo",
                    key=(
                        f"delete_photo_"
                        f"{original_index}"
                    ),
                ):

                    del (
                        st.session_state
                        .daily_photos[
                            original_index
                        ]
                    )

                    st.rerun()


    # ========================================================
    # 10 NOTES
    # ========================================================

    st.divider()

    st.subheader(
        "10. General Notes"
    )


    daily_notes = (
        st.text_area(
            "Daily Report Notes",
            value=daily_saved_record.get(
                "Notes",
                "",
            ),
            placeholder=(
                "General notes, coordination items, "
                "comments or information for the day."
            ),
            key=(
                f"daily_notes_"
                f"{daily_suffix}"
            ),
        )
    )


    # ========================================================
    # SUMMARY
    # ========================================================

    st.divider()

    st.header(
        "Daily Report Summary"
    )


    total_manpower = (
        calculate_total_manpower(
            daily_working_key
        )
    )


    total_plant = (
        calculate_total_equipment(
            daily_working_key
        )
    )


    activity_count = (
        len(
            current_activities
        )
    )


    inspection_count = (
        len(
            current_inspections
        )
    )


    issue_count = (
        len(
            current_issues
        )
    )


    photo_count = (
        len(
            current_photos
        )
    )


    locations_active = set()


    for item in (
        current_activities
    ):

        location = (
            item.get(
                "Location / Plot",
                "",
            )
            .strip()
        )

        if location:

            locations_active.add(
                location
            )


    sr1, sr2, sr3, sr4 = (
        st.columns(4)
    )

    with sr1:

        st.metric(
            "Total Manpower",
            total_manpower,
        )

    with sr2:

        st.metric(
            "Plant / Equipment",
            total_plant,
        )

    with sr3:

        st.metric(
            "Activities",
            activity_count,
        )

    with sr4:

        st.metric(
            "Active Locations",
            len(
                locations_active
            ),
        )


    sr5, sr6, sr7, sr8 = (
        st.columns(4)
    )

    with sr5:

        st.metric(
            "Deliveries",
            len(
                current_deliveries
            ),
        )

    with sr6:

        st.metric(
            "Inspections",
            inspection_count,
        )

    with sr7:

        st.metric(
            "Issues",
            issue_count,
        )

    with sr8:

        st.metric(
            "Photos",
            photo_count,
        )


    # ========================================================
    # SAVE / COMPLETE
    # ========================================================

    st.divider()


    save_col1, save_col2 = (
        st.columns(2)
    )

    with save_col1:

        save_draft = (
            st.button(
                "💾 Save as Draft",
                type="secondary",
                use_container_width=True,
                key=(
                    f"save_daily_draft_"
                    f"{daily_suffix}"
                ),
            )
        )

    with save_col2:

        complete_report = (
            st.button(
                "✅ Complete Daily Report",
                type="primary",
                use_container_width=True,
                key=(
                    f"complete_daily_"
                    f"{daily_suffix}"
                ),
            )
        )


    # ========================================================
    # SAVE LOGIC
    # ========================================================

    if (
        save_draft
        or
        complete_report
    ):

        if (
            not daily_project_input.strip()
        ):

            st.error(
                "Please enter Project / Site."
            )

        else:

            final_daily_key = (
                daily_report_key(
                    daily_project_input,
                    daily_date_input,
                )
            )


            duplicate = (
                daily_mode
                == "➕ New Daily Report"
                and
                final_daily_key
                in st.session_state
                .daily_reports
            )


            if duplicate:

                st.error(
                    (
                        "A Daily Report already exists "
                        "for this Project and Date. "
                        "Use Open Existing Daily Report."
                    )
                )

            else:

                report_status = (
                    "Complete"
                    if complete_report
                    else "Draft"
                )


                created_date = (
                    daily_saved_record.get(
                        "Created",
                        date.today(),
                    )
                )


                report_record = {
                    "Created":
                        created_date,

                    "Last Updated":
                        date.today(),

                    "Project":
                        daily_project_input,

                    "Date":
                        daily_date_input,

                    "Working Start":
                        working_start,

                    "Working Finish":
                        working_finish,

                    "Weather Morning":
                        weather_morning,

                    "Workability Morning":
                        workability_morning,

                    "Weather Afternoon":
                        weather_afternoon,

                    "Workability Afternoon":
                        workability_afternoon,

                    "Weather Notes":
                        weather_notes,

                    "Notes":
                        daily_notes,

                    "Status":
                        report_status,
                }


                st.session_state.daily_reports[
                    final_daily_key
                ] = report_record


                if (
                    daily_mode
                    == "➕ New Daily Report"
                    and
                    daily_working_key
                    != final_daily_key
                ):

                    daily_collections = [
                        st.session_state.daily_manpower,
                        st.session_state.daily_equipment,
                        st.session_state.daily_activities,
                        st.session_state.daily_deliveries,
                        st.session_state.daily_inspections,
                        st.session_state.daily_issues,
                        st.session_state.daily_photos,
                    ]


                    for collection in (
                        daily_collections
                    ):

                        for item in (
                            collection
                        ):

                            if (
                                item.get(
                                    "Daily Report Key"
                                )
                                == daily_working_key
                            ):

                                item[
                                    "Daily Report Key"
                                ] = final_daily_key


                if complete_report:

                    st.success(
                        "Daily Report completed."
                    )

                else:

                    st.success(
                        "Daily Report saved as Draft."
                    )


                st.rerun()


    # ========================================================
    # DOWNLOAD REPORTS
    # ========================================================

    report_available = (
        daily_working_key
        in st.session_state
        .daily_reports
    )


    if report_available:

        st.divider()

        st.header(
            "Report Export"
        )

        st.caption(
            "Generate a formatted Glenveagh report "
            "using the data recorded above."
        )


        export_record = (
            st.session_state
            .daily_reports[
                daily_working_key
            ]
        )


        export_project = (
            str(
                export_record.get(
                    "Project",
                    "Project",
                )
            )
            .replace(
                " ",
                "_",
            )
            .replace(
                "/",
                "-",
            )
        )


        export_date_value = (
            export_record.get(
                "Date",
                date.today(),
            )
        )


        if isinstance(
            export_date_value,
            date,
        ):

            export_date_name = (
                export_date_value.strftime(
                    "%Y-%m-%d"
                )
            )

        else:

            export_date_name = (
                str(
                    export_date_value
                )
            )


        try:

            pdf_bytes = (
                generate_daily_pdf(
                    daily_working_key
                )
            )


            excel_bytes = (
                generate_daily_excel(
                    daily_working_key
                )
            )


            download1, download2 = (
                st.columns(2)
            )


            with download1:

                st.download_button(
                    label="📄 Download PDF Report",
                    data=pdf_bytes,
                    file_name=(
                        f"Daily_Report_"
                        f"{export_project}_"
                        f"{export_date_name}.pdf"
                    ),
                    mime="application/pdf",
                    type="primary",
                    width="stretch",
                    key=(
                        f"download_pdf_"
                        f"{daily_suffix}"
                    ),
                )


            with download2:

                st.download_button(
                    label="📊 Download Excel Report",
                    data=excel_bytes,
                    file_name=(
                        f"Daily_Report_"
                        f"{export_project}_"
                        f"{export_date_name}.xlsx"
                    ),
                    mime=(
                        "application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet"
                    ),
                    width="stretch",
                    key=(
                        f"download_excel_"
                        f"{daily_suffix}"
                    ),
                )


        except Exception as export_error:

            st.error(
                (
                    "Could not generate the report: "
                    f"{export_error}"
                )
            )


    else:

        st.info(
            "Save the Daily Report first to enable PDF and Excel export."
        )


    # ========================================================
    # CURRENT STATUS
    # ========================================================

    if (
        daily_mode
        == "📂 Open Existing Daily Report"
    ):

        st.divider()


        current_report_record = (
            st.session_state
            .daily_reports[
                daily_current_key
            ]
        )


        current_status = (
            current_report_record.get(
                "Status",
                "Draft",
            )
        )


        if (
            current_status
            == "Complete"
        ):

            st.success(
                "✅ DAILY REPORT STATUS: COMPLETE"
            )

        else:

            st.info(
                "🟡 DAILY REPORT STATUS: DRAFT"
            )